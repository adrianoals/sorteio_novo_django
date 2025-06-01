from django.shortcuts import render, redirect
from django.utils import timezone
import random
from django.http import HttpResponse
from django.urls import reverse
from openpyxl import load_workbook
from io import BytesIO
from .models import Apartamento, Vaga, Sorteio
import time
import logging

logger = logging.getLogger(__name__)

# Create your views here.

# View para realizar o sorteio
def splendore_sorteio(request):
    if request.method == 'POST':
        start_time = time.time()
        
        # Limpar registros anteriores de sorteio
        logger.info("Iniciando deleção dos registros anteriores")
        delete_start = time.time()
        Sorteio.objects.all().delete()
        logger.info(f"Deleção concluída em {time.time() - delete_start:.2f} segundos")

        # Obter todos os apartamentos e vagas
        logger.info("Buscando apartamentos e vagas")
        fetch_start = time.time()
        apartamentos = list(Apartamento.objects.all())
        vagas = list(Vaga.objects.all())
        logger.info(f"Busca concluída em {time.time() - fetch_start:.2f} segundos")

        # Agrupar apartamentos e vagas por torre
        apartamentos_por_torre = {}
        vagas_por_torre = {}
        
        for apartamento in apartamentos:
            torre = apartamento.torre
            if torre not in apartamentos_por_torre:
                apartamentos_por_torre[torre] = []
            apartamentos_por_torre[torre].append(apartamento)
            
        for vaga in vagas:
            torre = vaga.torre
            if torre not in vagas_por_torre:
                vagas_por_torre[torre] = []
            vagas_por_torre[torre].append(vaga)

        # Lista para armazenar todos os objetos Sorteio
        sorteios_para_criar = []

        # Realizar o sorteio para cada torre
        logger.info("Iniciando criação dos registros de sorteio")
        create_start = time.time()
        
        for torre in range(1, 6):  # Torres de 1 a 5
            torre_str = str(torre)
            if torre_str in apartamentos_por_torre and torre_str in vagas_por_torre:
                # Embaralhar apartamentos e vagas da mesma torre
                apartamentos_torre = apartamentos_por_torre[torre_str]
                vagas_torre = vagas_por_torre[torre_str]
                
                random.shuffle(apartamentos_torre)
                random.shuffle(vagas_torre)

                # Alocar vagas para apartamentos da mesma torre
                for i in range(min(len(apartamentos_torre), len(vagas_torre))):
                    sorteios_para_criar.append(
                        Sorteio(
                            apartamento=apartamentos_torre[i],
                            vaga=vagas_torre[i]
                        )
                    )

        # Criar todos os registros de uma vez usando bulk_create
        Sorteio.objects.bulk_create(sorteios_para_criar)
        logger.info(f"Criação dos registros concluída em {time.time() - create_start:.2f} segundos")

        # Marcar o sorteio como iniciado e armazenar o horário de conclusão
        request.session['sorteio_iniciado'] = True
        request.session['horario_conclusao'] = timezone.localtime().strftime("%d/%m/%Y às %Hh %Mmin e %Ss")

        logger.info(f"Sorteio total concluído em {time.time() - start_time:.2f} segundos")
        
        # Redireciona para a mesma página após o sorteio
        return redirect(reverse('splendore_sorteio'))

    # Se o método for GET, exibe os resultados ou a interface de sorteio
    sorteio_iniciado = request.session.get('sorteio_iniciado', False)
    vagas_atribuidas = Sorteio.objects.exists()
    resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id') if vagas_atribuidas else None

    context = {
        'sorteio_iniciado': sorteio_iniciado,
        'vagas_atribuidas': vagas_atribuidas,
        'resultados_sorteio': resultados_sorteio,
        'horario_conclusao': request.session.get('horario_conclusao', '')
    }

    return render(request, 'splendore/splendore_sorteio.html', context)

def splendore_excel(request):
    # Caminho do modelo Excel
    caminho_modelo = 'setup/static/assets/modelos/sorteiosplendore.xlsx'

    # Carregar o modelo existente
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

    # Pegar o horário de conclusão do sorteio
    horario_conclusao = request.session.get('horario_conclusao', 'Horário não disponível')
    ws['A8'] = f"Sorteio realizado em: {horario_conclusao}"

    # Começar a partir da linha 10 (baseado no layout do seu modelo)
    linha = 10
    for sorteio in resultados_sorteio:
        ws[f'A{linha}'] = sorteio.apartamento.torre
        ws[f'B{linha}'] = sorteio.apartamento.numero
        ws[f'C{linha}'] = sorteio.vaga.numero
        ws[f'D{linha}'] = sorteio.vaga.torre
        linha += 1

    # Configurar a resposta para o download do Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sorteio_splendore.xlsx"'

    # Salvar o arquivo Excel na resposta
    wb.save(response)

    return response

def splendore_qrcode(request):
    # Obter todos os apartamentos para preencher o dropdown
    apartamentos_disponiveis = Apartamento.objects.all().order_by('torre', 'numero')
    
    # Obter o apartamento e torre selecionados através do filtro (via GET)
    numero_apartamento = request.GET.get('apartamento')
    torre_selecionada = request.GET.get('torre')

    # Inicializar a variável de resultados filtrados como uma lista vazia
    resultados_filtrados = []

    # Se o apartamento foi selecionado, realizar a filtragem dos resultados do sorteio
    if numero_apartamento:
        # Construir o filtro base com o número do apartamento
        filtro = {'apartamento__numero': numero_apartamento}
        
        # Se uma torre específica foi selecionada, adicionar ao filtro
        if torre_selecionada:
            filtro['apartamento__torre'] = torre_selecionada
            
        # Buscar os sorteios com os filtros aplicados
        resultados_filtrados = Sorteio.objects.filter(**filtro)

    return render(request, 'splendore/splendore_qrcode.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': numero_apartamento,
        'torre_selecionada': torre_selecionada,
        'apartamentos_disponiveis': apartamentos_disponiveis,
    })

def splendore_zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        return redirect('splendore_sorteio')
    else:
        return render(request, 'splendore/splendore_zerar.html')
