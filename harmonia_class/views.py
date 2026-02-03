import os
from django.shortcuts import render, redirect
from django.utils import timezone
from django.conf import settings
import random
from django.http import HttpResponse
from django.urls import reverse
from openpyxl import load_workbook
from openpyxl import Workbook
from .models import Apartamento, Vaga, Sorteio


def harmonia_class_sorteio(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()

        todos_apartamentos = list(Apartamento.objects.all())
        todas_vagas_simples = list(Vaga.objects.filter(tipo_vaga='Simples'))
        todas_vagas_duplas = list(Vaga.objects.filter(tipo_vaga='Dupla'))
        todas_vagas_pne = list(Vaga.objects.filter(is_pne=True))

        vagas_atribuidas_ids = set()
        apartamentos_atribuidos_ids = set()
        sorteios_para_criar = []

        # 1. PNE: 3 apartamentos PNE -> 3 vagas PNE
        apts_pne = [
            apt for apt in todos_apartamentos
            if apt.is_pne and apt.id not in apartamentos_atribuidos_ids
        ]
        random.shuffle(apts_pne)
        vagas_pne_disponiveis = [v for v in todas_vagas_pne if v.id not in vagas_atribuidas_ids]
        random.shuffle(vagas_pne_disponiveis)
        for i, apt in enumerate(apts_pne):
            if i < len(vagas_pne_disponiveis):
                vaga = vagas_pne_disponiveis[i]
                sorteios_para_criar.append(Sorteio(apartamento=apt, vaga=vaga))
                vagas_atribuidas_ids.add(vaga.id)
                apartamentos_atribuidos_ids.add(apt.id)

        # 2. Vagas duplas: 15 apartamentos -> 15 vagas duplas
        apts_vaga_dupla = [
            apt for apt in todos_apartamentos
            if apt.direito_vaga_dupla and apt.id not in apartamentos_atribuidos_ids
        ]
        random.shuffle(apts_vaga_dupla)
        vagas_duplas_disponiveis = [
            v for v in todas_vagas_duplas
            if v.id not in vagas_atribuidas_ids
        ]
        random.shuffle(vagas_duplas_disponiveis)
        for i, apt in enumerate(apts_vaga_dupla):
            if i < len(vagas_duplas_disponiveis):
                vaga = vagas_duplas_disponiveis[i]
                sorteios_para_criar.append(Sorteio(apartamento=apt, vaga=vaga))
                vagas_atribuidas_ids.add(vaga.id)
                apartamentos_atribuidos_ids.add(apt.id)

        # 3. Demais: 62 apartamentos -> 62 vagas simples
        apts_restantes = [
            apt for apt in todos_apartamentos
            if apt.id not in apartamentos_atribuidos_ids
        ]
        random.shuffle(apts_restantes)
        vagas_simples_disponiveis = [
            v for v in todas_vagas_simples
            if v.id not in vagas_atribuidas_ids
        ]
        random.shuffle(vagas_simples_disponiveis)
        for i, apt in enumerate(apts_restantes):
            if i < len(vagas_simples_disponiveis):
                vaga = vagas_simples_disponiveis[i]
                sorteios_para_criar.append(Sorteio(apartamento=apt, vaga=vaga))
                vagas_atribuidas_ids.add(vaga.id)
                apartamentos_atribuidos_ids.add(apt.id)

        if sorteios_para_criar:
            Sorteio.objects.bulk_create(sorteios_para_criar)

        request.session['harmonia_class_sorteio_iniciado'] = True
        request.session['harmonia_class_horario_conclusao'] = timezone.localtime().strftime(
            "%d/%m/%Y às %Hh %Mmin e %Ss"
        )
        return redirect(reverse('harmonia_class_sorteio'))

    sorteio_iniciado = request.session.get('harmonia_class_sorteio_iniciado', False)
    vagas_atribuidas = Sorteio.objects.exists()
    resultados_sorteio = Sorteio.objects.order_by('apartamento__id') if vagas_atribuidas else None

    context = {
        'sorteio_iniciado': sorteio_iniciado,
        'vagas_atribuidas': vagas_atribuidas,
        'resultados_sorteio': resultados_sorteio,
        'horario_conclusao': request.session.get('harmonia_class_horario_conclusao', ''),
    }
    return render(request, 'harmonia_class/harmonia_class_sorteio.html', context)


def harmonia_class_excel(request):
    resultados_sorteio = Sorteio.objects.select_related(
        'apartamento', 'vaga'
    ).order_by('apartamento__numero')
    horario_conclusao = request.session.get(
        'harmonia_class_horario_conclusao', 'Horário não disponível'
    )

    caminho_modelo = os.path.join(
        str(settings.BASE_DIR),
        'setup', 'static', 'assets', 'modelos', 'sorteioharmoniaclass.xlsx'
    )

    if os.path.isfile(caminho_modelo):
        wb = load_workbook(caminho_modelo)
        ws = wb.active
        ws['A8'] = f"Sorteio realizado em: {horario_conclusao}"
        linha_inicial = 10
        linha_maxima = ws.max_row
        for linha in range(linha_inicial, linha_maxima + 1):
            ws[f'A{linha}'] = None
            ws[f'B{linha}'] = None
            ws[f'C{linha}'] = None
            ws[f'D{linha}'] = None
        linha = 10
        for sorteio in resultados_sorteio:
            ws[f'A{linha}'] = sorteio.apartamento.numero
            ws[f'B{linha}'] = sorteio.vaga.numero
            ws[f'C{linha}'] = sorteio.vaga.localizacao
            ws[f'D{linha}'] = sorteio.vaga.tipo_vaga
            linha += 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sorteio Harmonia Class"
        ws['A1'] = "Sorteio - Condomínio Harmonia Class"
        ws['A2'] = f"Sorteio realizado em: {horario_conclusao}"
        ws['A4'] = "Apartamento"
        ws['B4'] = "Vaga"
        ws['C4'] = "Localização"
        ws['D4'] = "Tipo"
        linha = 5
        for sorteio in resultados_sorteio:
            ws[f'A{linha}'] = sorteio.apartamento.numero
            ws[f'B{linha}'] = sorteio.vaga.numero
            ws[f'C{linha}'] = sorteio.vaga.localizacao
            ws[f'D{linha}'] = sorteio.vaga.tipo_vaga
            linha += 1

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sorteio_harmonia_class.xlsx"'
    wb.save(response)
    return response


def harmonia_class_qrcode(request):
    apartamentos_disponiveis = Apartamento.objects.all()
    numero_apartamento = request.GET.get('apartamento')
    resultados_filtrados = []
    if numero_apartamento:
        resultados_filtrados = Sorteio.objects.filter(
            apartamento__numero=numero_apartamento
        )
    return render(request, 'harmonia_class/harmonia_class_qrcode.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': numero_apartamento,
        'apartamentos_disponiveis': apartamentos_disponiveis,
    })


def harmonia_class_zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        return redirect('harmonia_class_sorteio')
    return render(request, 'harmonia_class/harmonia_class_zerar.html')
