from django.shortcuts import render, redirect
from django.utils import timezone
import random
from django.http import HttpResponse
from django.urls import reverse
# from openpyxl import Workbook  # Para gerar o Excel
from openpyxl import load_workbook
from io import BytesIO  # Para manipular imagens em memória
from passaros.models import Apartamento,Bloco, Vaga, Sorteio


# def passaros_sorteio(request):
#     if request.method == 'POST':
#         # Marcar o sorteio como iniciado
#         request.session['sorteio_iniciado'] = True

#         # Redirecionar imediatamente para renderizar o estado do sorteio
#         return redirect(reverse('passaros_sorteio'))

#     # Sorteio sendo iniciado
#     sorteio_iniciado = request.session.get('sorteio_iniciado', False)

#     if sorteio_iniciado:
#         # Limpar registros anteriores de sorteio
#         Sorteio.objects.all().delete()

#         # Obter todos os apartamentos e vagas disponíveis
#         apartamentos = list(Apartamento.objects.all())
#         vagas_disponiveis = list(Vaga.objects.all())

#         # Embaralhar as listas para garantir aleatoriedade
#         random.shuffle(apartamentos)
#         random.shuffle(vagas_disponiveis)

#         # Realizar o sorteio
#         for apartamento in apartamentos:
#             # Filtrar vagas disponíveis apenas para o bloco do apartamento
#             vagas_bloco = [vaga for vaga in vagas_disponiveis if vaga.bloco == apartamento.bloco]

#             if vagas_bloco:
#                 # Seleciona a primeira vaga disponível no bloco
#                 vaga = vagas_bloco.pop(0)
#                 Sorteio.objects.create(apartamento=apartamento, vaga=vaga)
#                 vagas_disponiveis.remove(vaga)  # Remove a vaga da lista geral
#             else:
#                 # Não há vagas disponíveis no bloco do apartamento
#                 continue

#         # Marcar o sorteio como concluído
#         request.session['sorteio_iniciado'] = False
#         request.session['horario_conclusao'] = timezone.localtime().strftime("%d/%m/%Y às %Hh %Mmin e %Ss")

#     # Exibir os resultados do sorteio ordenados por ID do apartamento
#     vagas_atribuidas = Sorteio.objects.exists()
#     resultados_sorteio = Sorteio.objects.order_by('apartamento__id') if vagas_atribuidas else None

#     context = {
#         'sorteio_iniciado': sorteio_iniciado,
#         'vagas_atribuidas': vagas_atribuidas,
#         'resultados_sorteio': resultados_sorteio,
#         'horario_conclusao': request.session.get('horario_conclusao', ''),  # Exibe o horário de conclusão
#     }

#     return render(request, 'passaros/passaros_sorteio.html', context)

def passaros_sorteio(request):
    if request.method == 'POST':
        # Marcar o sorteio como iniciado
        request.session['sorteio_iniciado'] = True

        # Redirecionar imediatamente para renderizar o estado do sorteio
        return redirect(reverse('passaros_sorteio'))

    # Sorteio sendo iniciado
    sorteio_iniciado = request.session.get('sorteio_iniciado', False)

    if sorteio_iniciado:
        # Limpar registros anteriores de sorteio
        Sorteio.objects.all().delete()

        # Obter todos os apartamentos e vagas disponíveis
        apartamentos = list(Apartamento.objects.all())
        vagas_disponiveis = list(Vaga.objects.all())

         # Pré-definir vagas para alguns apartamentos
        apartamentos_com_vaga_predefinida = [
                (545, 681), 
                (277, 837), 
                (552, 690), 
                (211, 967), 
                (459, 903), 
                (544, 665), 
                (60, 1056), 
                # VAGAS PNE
                (171, 1294),
                (202, 868),
                (281, 834),
                (309, 828),
                (327, 821),
                (367, 808),
                (436, 788),
                (498, 764),
                (520, 676),
                (533, 684),
                (567, 696),
                (569, 697),
                (579, 704),
                (582, 705),
        ]

        # Criar sorteios pré-definidos
        for apartamento_id, vaga_id in apartamentos_com_vaga_predefinida:
            apartamento = Apartamento.objects.get(id=apartamento_id)
            vaga = Vaga.objects.get(id=vaga_id)
            Sorteio.objects.create(apartamento=apartamento, vaga=vaga)

            # Remover o apartamento e a vaga das listas
            if apartamento in apartamentos:
                apartamentos.remove(apartamento)
            if vaga in vagas_disponiveis:
                vagas_disponiveis.remove(vaga)

        # Embaralhar as listas restantes para garantir aleatoriedade
        random.shuffle(apartamentos)
        random.shuffle(vagas_disponiveis)

        # Realizar o sorteio para os apartamentos restantes
        for apartamento in apartamentos:
            # Filtrar vagas disponíveis apenas para o bloco do apartamento
            vagas_bloco = [vaga for vaga in vagas_disponiveis if vaga.bloco == apartamento.bloco]

            if vagas_bloco:
                # Seleciona a primeira vaga disponível no bloco
                vaga = vagas_bloco.pop(0)
                Sorteio.objects.create(apartamento=apartamento, vaga=vaga)
                vagas_disponiveis.remove(vaga)  # Remove a vaga da lista geral
            else:
                # Não há vagas disponíveis no bloco do apartamento
                continue

        # Marcar o sorteio como concluído
        request.session['sorteio_iniciado'] = False
        request.session['horario_conclusao'] = timezone.localtime().strftime("%d/%m/%Y às %Hh %Mmin e %Ss")

    # Exibir os resultados do sorteio ordenados por ID do apartamento
    vagas_atribuidas = Sorteio.objects.exists()
    resultados_sorteio = Sorteio.objects.order_by('apartamento__id') if vagas_atribuidas else None

    context = {
        'sorteio_iniciado': sorteio_iniciado,
        'vagas_atribuidas': vagas_atribuidas,
        'resultados_sorteio': resultados_sorteio,
        'horario_conclusao': request.session.get('horario_conclusao', ''),  # Exibe o horário de conclusão
    }

    return render(request, 'passaros/passaros_sorteio.html', context)



def passaros_excel(request):
    # Caminho do modelo Excel
    caminho_modelo = 'setup/static/assets/modelos/sorteiopassaros.xlsx'

    # Carregar o modelo existente
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

    # Pegar o horário de conclusão do sorteio
    horario_conclusao = request.session.get('horario_conclusao', 'Horário não disponível')
    ws['A8'] = f"Sorteio realizado em: {horario_conclusao}"

    # Começar a partir da linha 10 (baseado no layout do seu modelo)
    linha = 10
    for sorteio in resultados_sorteio:
        ws[f'A{linha}'] = f"Bloco {sorteio.vaga.bloco.numero}" if sorteio.vaga.bloco else "Sem Bloco"
        ws[f'B{linha}'] = f'Unidade {sorteio.apartamento.numero}'  # Número do apartamento
        ws[f'C{linha}'] = sorteio.vaga.numero  # Número da vaga
        linha += 1

    # Configurar a resposta para o download do Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sorteio_passaros.xlsx"'

    # Salvar o arquivo Excel na resposta
    wb.save(response)

    return response


def passaros_qrcode(request):
    # Obter todos os blocos disponíveis
    blocos_disponiveis = Bloco.objects.all()

    # Obter o bloco selecionado através do filtro (via GET)
    numero_bloco = request.GET.get('bloco')

    # Inicializar apartamentos e resultados filtrados
    apartamentos_disponiveis = []
    resultados_filtrados = []

    # Se um bloco foi selecionado, filtrar os apartamentos disponíveis naquele bloco
    if numero_bloco:
        apartamentos_disponiveis = Apartamento.objects.filter(bloco__numero=numero_bloco)

        # Obter o apartamento selecionado através do filtro
        numero_apartamento = request.GET.get('apartamento')

        if numero_apartamento:
            # Filtrar os resultados do sorteio para o apartamento selecionado
            resultados_filtrados = Sorteio.objects.filter(
                apartamento__numero=numero_apartamento,
                apartamento__bloco__numero=numero_bloco
            )

    else:
        numero_apartamento = None

    return render(request, 'passaros/passaros_qrcode.html', {
        'blocos_disponiveis': blocos_disponiveis,
        'apartamentos_disponiveis': apartamentos_disponiveis,
        'resultados_filtrados': resultados_filtrados,
        'bloco_selecionado': numero_bloco,
        'apartamento_selecionado': numero_apartamento,
    })



def passaros_zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        return redirect('passaros_sorteio')
    else:
        return render(request, 'passaros/passaros_zerar.html')
    

