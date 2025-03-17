from django.shortcuts import render, redirect
from .models import Apartamento, Vaga, Sorteio
from django.utils import timezone
import random
from django.contrib.admin.views.decorators import staff_member_required



@staff_member_required
def arthur_aleatorio(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio
        Sorteio.objects.all().delete()
        
        # Obter todos os apartamentos e grupos de vagas
        apartamentos = list(Apartamento.objects.all())
        vagas = list(Vaga.objects.all())

        # Certifique-se de que existem vagas suficientes para todos os apartamentos
        if len(vagas) >= len(apartamentos):
            random.shuffle(vagas)

            for apartamento in apartamentos:
                vaga_selecionada = vagas.pop()
                Sorteio.objects.create(
                    apartamento=apartamento, 
                    vaga=vaga_selecionada
                )
        else:
           
            pass
        
        # Armazenar informações do sorteio na sessão
        request.session['sorteio_iniciado_nc'] = True
        request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

        return redirect('arthur_aleatorio')
    
    else:
        sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
        resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
        vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

        return render(request, 'arthur/arthur_aleatorio.html', {
            'resultados_sorteio_nc': resultados_sorteio_nc,
            'vagas_atribuidas_nc': vagas_atribuidas_nc,
            'sorteio_iniciado_nc': sorteio_iniciado_nc,
            'horario_conclusao_nc': request.session.get('horario_conclusao_nc', '')
        })


@staff_member_required
def zerar_arthur(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        return redirect('arthur_aleatorio')
    else:
        return render(request, 'arthur/arthur_zerar.html')


# Excel
from openpyxl import load_workbook
from django.http import HttpResponse
from django.utils import timezone
from django.contrib import messages

def excel_arthur(request):
    caminho_modelo = 'static/assets/modelos/sorteioarthur.xlsx'

    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

    horario_conclusao_nc = request.session.get('horario_conclusao_nc', 'Horário não disponível')
    ws['A8'] = f"Sorteio realizado em: {horario_conclusao_nc}"

    linha = 11
    for sorteio in resultados_sorteio_nc:
        ws[f'A{linha}'] = sorteio.apartamento.numero_apartamento
        ws[f'B{linha}'] = sorteio.vaga.vaga
        linha += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nome_arquivo = "resultado_sorteio.xlsx"
    response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'

    wb.save(response)

    return response


def qrcode_arthur(request):
    # Obter todos os blocos disponíveis
    blocos_disponiveis = Bloco.objects.all().order_by('nome')

    # Obter o bloco selecionado através do filtro (via GET)
    nome_bloco = request.GET.get('bloco')
    numero_apartamento = request.GET.get('apartamento')

    # Inicializar apartamentos e resultados filtrados
    apartamentos_disponiveis = Apartamento.objects.filter(bloco__nome=nome_bloco) if nome_bloco else None
    resultados_filtrados = Sorteio.objects.filter(
        apartamento__numero_apartamento=numero_apartamento,
        apartamento__bloco__nome=nome_bloco
    ) if numero_apartamento else None

    return render(request, 'arthur/arthur_qrcode.html', {
        'blocos_disponiveis': blocos_disponiveis,
        'apartamentos_disponiveis': apartamentos_disponiveis,
        'resultados_filtrados': resultados_filtrados,
        'bloco_selecionado': nome_bloco,
        'apartamento_selecionado': numero_apartamento,
    })



@staff_member_required
def arthur_presenca(request):
    if request.method == 'POST':
        apartamento = Apartamento.objects.all()
        for item in apartamento:
            # Atualiza o campo de presença
            item.presenca = request.POST.get('presenca' + str(item.id)) == 'True'
            # Atualiza o campo de prioridade
            item.prioridade = request.POST.get('prioridade' + str(item.id)) == 'True'
            item.save()
        return redirect('arthur_filtrar')  # Redireciona para a rota 'filtrar_presenca'

    apartamento = Apartamento.objects.all()
    return render(request, 'arthur/arthur_presenca.html', {"lista_de_presenca": apartamento})



@staff_member_required
def arthur_filtrar(request):
    lista_de_presenca = Apartamento.objects.none()  # Inicia com uma queryset vazia
    if request.method == 'POST':
        lista_de_presenca = Apartamento.objects.all()  # Recupera todos os objetos quando o formulário é submetido
        if 'presentes' in request.POST:
            lista_de_presenca = lista_de_presenca.filter(presenca=True)
        if 'ausentes' in request.POST:
            lista_de_presenca = lista_de_presenca.filter(presenca=False)
    return render(request, 'arthur/arthur_filtrar.html', {"lista_de_presenca": lista_de_presenca})

import random
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from .models import Apartamento, Vaga, Sorteio

@staff_member_required
def arthur_s_apartamento(request):
    # Filtra apartamentos disponíveis (prioritários primeiro)
    apartamentos_prioritarios = Apartamento.objects.filter(presenca=True, prioridade=True).exclude(sorteio__isnull=False)
    apartamentos_nao_prioritarios = Apartamento.objects.filter(presenca=True, prioridade=False).exclude(sorteio__isnull=False)

    # Verifica se o sorteio foi finalizado
    sorteio_finalizado = not (apartamentos_prioritarios.exists() or apartamentos_nao_prioritarios.exists())
    apartamento_sorteado = None

    if request.method == 'POST':
        if 'realizar_sorteio' in request.POST:
            # Tenta sortear primeiro entre os prioritários, se não houver, entre os não prioritários
            if apartamentos_prioritarios.exists():
                apartamento_sorteado = random.choice(list(apartamentos_prioritarios))
            elif apartamentos_nao_prioritarios.exists():
                apartamento_sorteado = random.choice(list(apartamentos_nao_prioritarios))
            else:
                messages.error(request, 'Nenhum apartamento disponível para sorteio.')
                return redirect('arthur_s_apartamento')

            messages.success(request, f'Apartamento {apartamento_sorteado.numero_apartamento} foi selecionado para atribuição de vaga.')

            # Lista todas as vagas disponíveis (sem restrições de bloco)
            vagas_disponiveis = Vaga.objects.filter(sorteio__isnull=True)

            return render(request, 'arthur/arthur_s_apartamento.html', {
                'sorteio_finalizado': sorteio_finalizado,
                'apartamentos_disponiveis': apartamentos_prioritarios.union(apartamentos_nao_prioritarios),
                'vagas_disponiveis': vagas_disponiveis,
                'item_de_presenca': apartamento_sorteado
            })

        # Associação da vaga ao apartamento sorteado
        if 'vaga_selecionada' in request.POST and 'apartamento_id' in request.POST:
            apartamento_id = request.POST.get('apartamento_id')
            vaga_selecionada = request.POST.get('vaga_selecionada')

            apartamento = Apartamento.objects.get(id=apartamento_id)
            vaga = Vaga.objects.get(vaga=vaga_selecionada)

            # Cria o sorteio e associa o apartamento à vaga
            Sorteio.objects.create(apartamento=apartamento, vaga=vaga)

            messages.success(request, f'Vaga {vaga.vaga} foi confirmada para o apartamento {apartamento.numero_apartamento} com sucesso!')
            return redirect('arthur_s_apartamento')

    # Renderização inicial
    return render(request, 'arthur/arthur_s_apartamento.html', {
        'sorteio_finalizado': sorteio_finalizado,
        'apartamentos_disponiveis': apartamentos_prioritarios.union(apartamentos_nao_prioritarios),
        'vagas_disponiveis': Vaga.objects.filter(sorteio__isnull=True),
        'item_de_presenca': apartamento_sorteado
    })


import random
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from .models import Apartamento, Vaga, Sorteio

@staff_member_required
def arthur_final(request):
    if request.method == 'POST':
        # Remover registros de sorteio anteriores para apartamentos ausentes (presença=False)
        Sorteio.objects.filter(apartamento__presenca=False).delete()

        # Filtrar apartamentos ausentes e vagas ainda não sorteadas
        apartamentos_ausentes = list(Apartamento.objects.filter(presenca=False).exclude(sorteio__isnull=False))
        vagas_disponiveis = list(Vaga.objects.exclude(id__in=Sorteio.objects.values_list('vaga_id', flat=True)))

        # Verifica se há vagas disponíveis para os apartamentos restantes
        if not vagas_disponiveis:
            messages.error(request, "Não há vagas disponíveis para sorteio.")
            return redirect('arthur_final')

        # Embaralha as vagas para distribuição aleatória
        random.shuffle(vagas_disponiveis)

        # Sorteio dos apartamentos ausentes
        for apartamento in apartamentos_ausentes:
            if vagas_disponiveis:
                vaga_selecionada = vagas_disponiveis.pop()
                Sorteio.objects.create(apartamento=apartamento, vaga=vaga_selecionada)
            else:
                messages.warning(request, f"Não há vagas suficientes para todos os apartamentos.")

        # Armazena na sessão que o sorteio foi concluído
        request.session['sorteio_iniciado_nc'] = True
        request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

        return redirect('arthur_final')

    else:
        # Verificar se o sorteio foi iniciado
        sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
        
        # Verificar a quantidade total de apartamentos e quantos já foram sorteados
        todos_apartamentos = Apartamento.objects.count()
        apartamentos_sorteados = Sorteio.objects.count()
        sorteio_completo = todos_apartamentos == apartamentos_sorteados

        # Obter os resultados do sorteio
        resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id')

        return render(request, 'arthur/arthur_final.html', {
            'resultados_sorteio_nc': resultados_sorteio_nc,
            'sorteio_iniciado_nc': sorteio_iniciado_nc,
            'horario_conclusao_nc': request.session.get('horario_conclusao_nc', ''),
            'sorteio_completo': sorteio_completo
        })

