from django.shortcuts import render, redirect
from django.utils import timezone
import random
from django.http import HttpResponse
from django.urls import reverse
from openpyxl import load_workbook
from io import BytesIO
from .models import Apartamento, Vaga, Sorteio

# Create your views here.

# View para realizar o sorteio
def alvorada_sorteio(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio
        Sorteio.objects.all().delete()

        # Obter todos os apartamentos e vagas
        apartamentos = list(Apartamento.objects.all())
        vagas = list(Vaga.objects.all())

        # Separar apartamentos com direito a vaga dupla
        apartamentos_vaga_dupla = [apt for apt in apartamentos if apt.direito_vaga_dupla]
        apartamentos_vaga_simples = [apt for apt in apartamentos if not apt.direito_vaga_dupla]

        # Separar vagas duplas e simples
        vagas_duplas = [vaga for vaga in vagas if vaga.tipo_vaga == 'Dupla']
        vagas_simples = [vaga for vaga in vagas if vaga.tipo_vaga == 'Simples']

        resultados_sorteio = []

        # 1. Sorteio de apartamentos com direito a vaga dupla
        random.shuffle(apartamentos_vaga_dupla)
        random.shuffle(vagas_duplas)

        # Alocar vagas duplas para apartamentos com direito
        for i in range(min(len(apartamentos_vaga_dupla), len(vagas_duplas))):
            apartamento = apartamentos_vaga_dupla[i]
            vaga = vagas_duplas[i]
            sorteio = Sorteio.objects.create(apartamento=apartamento, vaga=vaga)
            resultados_sorteio.append(sorteio)
            vagas.remove(vaga)  # Remover vaga da lista total

        # 2. Sorteio dos apartamentos restantes com as vagas restantes
        # Juntar todos os apartamentos que ainda não receberam vaga
        apartamentos_restantes = apartamentos_vaga_dupla[len(vagas_duplas):] + apartamentos_vaga_simples
        vagas_restantes = vagas  # Todas as vagas que sobraram

        # Embaralhar as listas
        random.shuffle(apartamentos_restantes)
        random.shuffle(vagas_restantes)

        # Alocar vagas restantes para apartamentos restantes
        for i in range(min(len(apartamentos_restantes), len(vagas_restantes))):
            apartamento = apartamentos_restantes[i]
            vaga = vagas_restantes[i]
            sorteio = Sorteio.objects.create(apartamento=apartamento, vaga=vaga)
            resultados_sorteio.append(sorteio)

        # Marcar o sorteio como iniciado e armazenar o horário de conclusão
        request.session['sorteio_iniciado'] = True
        request.session['horario_conclusao'] = timezone.localtime().strftime("%d/%m/%Y às %Hh %Mmin e %Ss")

        # Redireciona para a mesma página após o sorteio
        return redirect(reverse('alvorada_sorteio'))

    # Se o método for GET, exibe os resultados ou a interface de sorteio
    sorteio_iniciado = request.session.get('sorteio_iniciado', False)
    vagas_atribuidas = Sorteio.objects.exists()
    resultados_sorteio = Sorteio.objects.order_by('apartamento__id') if vagas_atribuidas else None

    context = {
        'sorteio_iniciado': sorteio_iniciado,
        'vagas_atribuidas': vagas_atribuidas,
        'resultados_sorteio': resultados_sorteio,
        'horario_conclusao': request.session.get('horario_conclusao', '')
    }

    return render(request, 'alvorada/alvorada_sorteio.html', context)

def alvorada_excel(request):
    # Caminho do modelo Excel
    caminho_modelo = 'setup/static/assets/modelos/sorteioalvorada.xlsx'

    # Carregar o modelo existente
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

    # Pegar o horário de conclusão do sorteio
    horario_conclusao = request.session.get('horario_conclusao', 'Horário não disponível')
    ws['A8'] = f"Sorteio realizado em: {horario_conclusao}"

    # Começar a partir da linha 10 (baseado no layout do seu modelo)
    linha = 10
    for sorteio in resultados_sorteio:
        ws[f'A{linha}'] = sorteio.apartamento.numero
        ws[f'B{linha}'] = sorteio.vaga.numero
        ws[f'C{linha}'] = f'Subsolo {sorteio.vaga.subsolo}'
        ws[f'D{linha}'] = sorteio.vaga.tipo_vaga
        linha += 1

    # Configurar a resposta para o download do Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sorteio_alvorada.xlsx"'

    # Salvar o arquivo Excel na resposta
    wb.save(response)

    return response

def alvorada_qrcode(request):
    # Obter todos os apartamentos para preencher o dropdown
    apartamentos_disponiveis = Apartamento.objects.all()
    
    # Obter o apartamento selecionado através do filtro (via GET)
    numero_apartamento = request.GET.get('apartamento')

    # Inicializar a variável de resultados filtrados como uma lista vazia
    resultados_filtrados = []

    # Se o apartamento foi selecionado, realizar a filtragem dos resultados do sorteio
    if numero_apartamento:
        # Buscar os sorteios para o apartamento selecionado
        resultados_filtrados = Sorteio.objects.filter(apartamento__numero=numero_apartamento)

    return render(request, 'alvorada/alvorada_qrcode.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': numero_apartamento,
        'apartamentos_disponiveis': apartamentos_disponiveis,
    })

def alvorada_zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        return redirect('alvorada_sorteio')
    else:
        return render(request, 'alvorada/alvorada_zerar.html')
