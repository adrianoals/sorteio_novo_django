from django.shortcuts import render, redirect
from .models import Apartamento, Vaga, Sorteio, Bloco
from django.utils import timezone
import random
from django.contrib.admin.views.decorators import staff_member_required

# Excel
from openpyxl import load_workbook
from django.http import HttpResponse
from django.utils import timezone
from django.contrib import messages


@staff_member_required
def ventura_aleatorio(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio
        Sorteio.objects.all().delete()
        
        # Obter todos os apartamentos e grupos de vagas
        apartamentos = list(Apartamento.objects.all())
        vagas = list(Vaga.objects.all())

        # Certifique-se de que existem vagas suficientes para todos os apartamentos
        if len(vagas) >= len(apartamentos):
            random.shuffle(vagas)

            for apartamento in apartamentos:
                vaga_selecionada = vagas.pop()
                Sorteio.objects.create(
                    apartamento=apartamento, 
                    vaga=vaga_selecionada
                )
        else:
           
            pass
        
        # Armazenar informações do sorteio na sessão
        request.session['sorteio_iniciado_nc'] = True
        request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

        return redirect('ventura_aleatorio')
    
    else:
        sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
        resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
        vagas_atribuidas_nc = resultados_sorteio_nc.exists()  # Verificar se existem resultados

        return render(request, 'ventura/ventura_aleatorio.html', {
            'resultados_sorteio_nc': resultados_sorteio_nc,
            'vagas_atribuidas_nc': vagas_atribuidas_nc,
            'sorteio_iniciado_nc': sorteio_iniciado_nc,
            'horario_conclusao_nc': request.session.get('horario_conclusao_nc', '')
        })


@staff_member_required
def zerar_ventura(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        return redirect('ventura_aleatorio')
    else:
        return render(request, 'ventura/ventura_zerar.html')


def excel_ventura(request):
    caminho_modelo = 'static/assets/modelos/sorteioventura.xlsx'

    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

    horario_conclusao_nc = request.session.get('horario_conclusao_nc', 'Horário não disponível')
    ws['A8'] = f"Sorteio realizado em: {horario_conclusao_nc}"

    linha = 10
    for sorteio in resultados_sorteio_nc:
        ws[f'A{linha}'] = sorteio.apartamento.bloco.nome
        ws[f'B{linha}'] = sorteio.apartamento.numero_apartamento
        ws[f'C{linha}'] = sorteio.vaga.vaga
        ws[f'D{linha}'] = sorteio.vaga.subsolo
        linha += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nome_arquivo = "resultado_sorteio.xlsx"
    response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'

    wb.save(response)

    return response



def qrcode_ventura(request):
    # Obter todos os blocos disponíveis
    blocos_disponiveis = Bloco.objects.all().order_by('nome')

    # Obter o bloco selecionado através do filtro (via GET)
    nome_bloco = request.GET.get('bloco')
    numero_apartamento = request.GET.get('apartamento')

    # Inicializar apartamentos e resultados filtrados
    apartamentos_disponiveis = Apartamento.objects.filter(bloco__nome=nome_bloco) if nome_bloco else None
    resultados_filtrados = Sorteio.objects.filter(
        apartamento__numero_apartamento=numero_apartamento,
        apartamento__bloco__nome=nome_bloco
    ) if numero_apartamento else None

    return render(request, 'ventura/ventura_qrcode.html', {
        'blocos_disponiveis': blocos_disponiveis,
        'apartamentos_disponiveis': apartamentos_disponiveis,
        'resultados_filtrados': resultados_filtrados,
        'bloco_selecionado': nome_bloco,
        'apartamento_selecionado': numero_apartamento,
    })


@staff_member_required
def ventura_presenca(request):
    if request.method == 'POST':
        apartamento = Apartamento.objects.all()
        for item in apartamento:
            # Atualiza o campo de presença
            item.presenca = request.POST.get('presenca' + str(item.id)) == 'True'
            # Atualiza o campo de prioridade
            item.prioridade = request.POST.get('prioridade' + str(item.id)) == 'True'
            item.save()
        return redirect('ventura_filtrar')  # Redireciona para a rota 'filtrar_presenca'

    apartamento = Apartamento.objects.all()
    return render(request, 'ventura/ventura_presenca.html', {"lista_de_presenca": apartamento})


@staff_member_required
def ventura_filtrar(request):
    lista_de_presenca = Apartamento.objects.none()  # Inicia com uma queryset vazia
    if request.method == 'POST':
        lista_de_presenca = Apartamento.objects.all()  # Recupera todos os objetos quando o formulário é submetido
        if 'presentes' in request.POST:
            lista_de_presenca = lista_de_presenca.filter(presenca=True)
        if 'ausentes' in request.POST:
            lista_de_presenca = lista_de_presenca.filter(presenca=False)
    return render(request, 'ventura/ventura_filtrar.html', {"lista_de_presenca": lista_de_presenca})



from django.shortcuts import render, redirect
from ventura.models import Apartamento, Vaga, Sorteio
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
import random

@staff_member_required
def ventura_s_apartamento(request):
    # Filtra apartamentos disponíveis
    apartamentos_prioritarios = Apartamento.objects.filter(presenca=True, prioridade=True).exclude(sorteio__isnull=False)
    apartamentos_nao_prioritarios = Apartamento.objects.filter(presenca=True, prioridade=False).exclude(sorteio__isnull=False)
    
    # Lista de IDs dos 12 apartamentos específicos
    apartamentos_especificos_ids = [223, 224, 229, 230, 235, 236, 345, 346, 351, 352, 357, 358]  
    vagas_especificas_ids = [351, 352, 353, 354, 355, 356, 357, 358, 359, 360, 361, 362] 
    
    # Verifica se o sorteio foi finalizado
    sorteio_finalizado = not (apartamentos_prioritarios.exists() or apartamentos_nao_prioritarios.exists())
    item_de_presenca = None

    if request.method == 'POST':
        if 'realizar_sorteio' in request.POST:
            if apartamentos_prioritarios.exists():
                # Sorteia entre os apartamentos prioritários
                apartamento_escolhido = random.choice(list(apartamentos_prioritarios))
            elif apartamentos_nao_prioritarios.exists():
                # Sorteia entre os apartamentos não prioritários
                apartamento_escolhido = random.choice(list(apartamentos_nao_prioritarios))
            else:
                messages.error(request, 'Nenhum apartamento disponível para sorteio.')
                return redirect('ventura_s_apartamento')

            messages.success(request, f'Apartamento {apartamento_escolhido.numero_apartamento} foi selecionado para atribuição de vaga.')

            # Verifica se o apartamento sorteado é um dos específicos
            if apartamento_escolhido.id in apartamentos_especificos_ids:
                # Filtra apenas as vagas específicas que ainda estão disponíveis
                vagas_disponiveis = Vaga.objects.filter(id__in=vagas_especificas_ids, sorteio__isnull=True)
            else:
                # Filtra vagas disponíveis do mesmo bloco que o apartamento sorteado
                vagas_disponiveis = Vaga.objects.filter(sorteio__isnull=True, bloco=apartamento_escolhido.bloco)

            # Renderiza a página com o apartamento sorteado e as vagas filtradas
            return render(request, 'ventura/ventura_s_apartamento.html', {
                'sorteio_finalizado': sorteio_finalizado,
                'apartamentos_disponiveis': apartamentos_prioritarios.union(apartamentos_nao_prioritarios),
                'vagas_disponiveis': vagas_disponiveis,
                'item_de_presenca': apartamento_escolhido
            })

        # Seção para associar a vaga ao apartamento sorteado
        if 'vaga_selecionada' in request.POST and 'apartamento_id' in request.POST:
            apartamento_id = request.POST.get('apartamento_id')
            vaga_selecionada = request.POST.get('vaga_selecionada')

            apartamento = Apartamento.objects.get(id=apartamento_id)
            vaga = Vaga.objects.get(vaga=vaga_selecionada)

            # Cria o sorteio associando o apartamento à vaga
            novo_sorteio = Sorteio(apartamento=apartamento, vaga=vaga)
            novo_sorteio.save()

            messages.success(request, f'Vaga {vaga.vaga} foi confirmada para o apartamento {apartamento.numero_apartamento} com sucesso!')
            return redirect('ventura_s_apartamento')

    # Renderização inicial
    return render(request, 'ventura/ventura_s_apartamento.html', {
        'sorteio_finalizado': sorteio_finalizado,
        'apartamentos_disponiveis': apartamentos_prioritarios.union(apartamentos_nao_prioritarios),
        'vagas_disponiveis': Vaga.objects.filter(sorteio__isnull=True),
        'item_de_presenca': item_de_presenca
    })


from django.shortcuts import render, redirect
from ventura.models import Apartamento, Vaga, Sorteio
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
import random

@staff_member_required
def ventura_final(request):
    # Lista de IDs dos apartamentos e vagas específicas
    apartamentos_especificos_ids = [223, 224, 229, 230, 235, 236, 345, 346, 351, 352, 357, 358]
    vagas_especificas_ids = [351, 352, 353, 354, 355, 356, 357, 358, 359, 360, 361, 362]

    if request.method == 'POST':
        # Limpar registros anteriores de sorteio para apartamentos com presença False
        Sorteio.objects.filter(apartamento__presenca=False).delete()

        # Filtrar apartamentos com presença False
        apartamentos = list(Apartamento.objects.filter(presenca=False))
        vagas_disponiveis = Vaga.objects.exclude(id__in=Sorteio.objects.values_list('vaga_id', flat=True))

        # Etapa 1: Atribuição para apartamentos específicos
        apartamentos_especificos = [apt for apt in apartamentos if apt.id in apartamentos_especificos_ids]
        for apartamento in apartamentos_especificos:
            vagas_disponiveis_especificas = list(vagas_disponiveis.filter(id__in=vagas_especificas_ids))
            if vagas_disponiveis_especificas:
                random.shuffle(vagas_disponiveis_especificas)
                vaga_selecionada = vagas_disponiveis_especificas.pop()
                Sorteio.objects.create(apartamento=apartamento, vaga=vaga_selecionada)
                vagas_disponiveis = vagas_disponiveis.exclude(id=vaga_selecionada.id)

        # Etapa 2: Atribuição para o restante dos apartamentos
        apartamentos_restantes = [apt for apt in apartamentos if apt.id not in apartamentos_especificos_ids]
        for apartamento in apartamentos_restantes:
            vagas_mesmo_bloco = list(vagas_disponiveis.filter(bloco=apartamento.bloco))
            if vagas_mesmo_bloco:
                random.shuffle(vagas_mesmo_bloco)
                vaga_selecionada = vagas_mesmo_bloco.pop()
                Sorteio.objects.create(apartamento=apartamento, vaga=vaga_selecionada)
                vagas_disponiveis = vagas_disponiveis.exclude(id=vaga_selecionada.id)
            else:
                messages.warning(request, f"Não há vagas disponíveis para o apartamento {apartamento.numero_apartamento} no bloco {apartamento.bloco}.")

        # Armazenar informações do sorteio na sessão
        request.session['sorteio_iniciado_nc'] = True
        request.session['horario_conclusao_nc'] = timezone.localtime().strftime("%d/%m/%Y às %Hh e %Mmin e %Ss")

        return redirect('ventura_final')

    else:
        sorteio_iniciado_nc = request.session.get('sorteio_iniciado_nc', False)
        todos_apartamentos = Apartamento.objects.count()
        apartamentos_sorteio = Sorteio.objects.count()
        vagas_atribuidas_completas = todos_apartamentos == apartamentos_sorteio

        resultados_sorteio_nc = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
        vagas_atribuidas_nc = resultados_sorteio_nc.exists()

        return render(request, 'ventura/ventura_final.html', {
            'resultados_sorteio_nc': resultados_sorteio_nc,
            'vagas_atribuidas_nc': vagas_atribuidas_nc,
            'sorteio_iniciado_nc': sorteio_iniciado_nc,
            'horario_conclusao_nc': request.session.get('horario_conclusao_nc', ''),
            'vagas_atribuidas_completas': vagas_atribuidas_completas
        })

