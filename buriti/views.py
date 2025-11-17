from django.shortcuts import render, redirect
from django.utils import timezone
import random
from django.http import HttpResponse
from django.urls import reverse
from openpyxl import load_workbook
from .models import Apartamento, Vaga, Sorteio

# View para realizar o sorteio
def buriti_sorteio(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio
        Sorteio.objects.all().delete()

        # Carregar todos os apartamentos e vagas
        todos_apartamentos = list(Apartamento.objects.all())
        todas_vagas = list(Vaga.objects.all())

        print(f"\n📊 DEBUG: Total de Apartamentos: {len(todos_apartamentos)} | Total de Vagas: {len(todas_vagas)}")
        
        # Verificar se os números estão corretos
        if len(todos_apartamentos) != 126:
            print(f"⚠️ ATENÇÃO: Esperado 126 apartamentos, encontrado {len(todos_apartamentos)}")
        if len(todas_vagas) != 126:
            print(f"⚠️ ATENÇÃO: Esperado 126 vagas, encontrado {len(todas_vagas)}")

        # Usar set para rastrear IDs de vagas e apartamentos já atribuídos
        vagas_atribuidas_ids = set()
        apartamentos_atribuidos_ids = set()
        sorteios_para_criar = []

        # ============================================
        # SORTEIO SIMPLES: 126 apartamentos para 126 vagas
        # ============================================
        random.shuffle(todos_apartamentos)
        random.shuffle(todas_vagas)
        
        # Atribuir uma vaga para cada apartamento
        for i, apt in enumerate(todos_apartamentos):
            if i < len(todas_vagas):
                vaga = todas_vagas[i]
                if vaga.id not in vagas_atribuidas_ids:
                    sorteios_para_criar.append(Sorteio(apartamento=apt, vaga=vaga))
                    vagas_atribuidas_ids.add(vaga.id)
                    apartamentos_atribuidos_ids.add(apt.id)
                    print(f"✅ {apt} → {vaga}")
            else:
                print(f"⚠️ {apt} não recebeu vaga (vagas insuficientes)")

        # ============================================
        # BULK CREATE (otimização de performance)
        # ============================================
        if sorteios_para_criar:
            Sorteio.objects.bulk_create(sorteios_para_criar)
            print(f"\n✅ Total de {len(sorteios_para_criar)} sorteios criados com sucesso!")

        # Marcar o sorteio como iniciado e armazenar o horário de conclusão
        request.session['sorteio_iniciado'] = True
        request.session['horario_conclusao'] = timezone.localtime().strftime("%d/%m/%Y às %Hh %Mmin e %Ss")

        # Redireciona para a mesma página após o sorteio
        return redirect(reverse('buriti_sorteio'))

    # Se o método for GET, exibe os resultados ou a interface de sorteio
    sorteio_iniciado = request.session.get('sorteio_iniciado', False)
    vagas_atribuidas = Sorteio.objects.exists()
    resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id') if vagas_atribuidas else None

    context = {
        'sorteio_iniciado': sorteio_iniciado,
        'vagas_atribuidas': vagas_atribuidas,
        'resultados_sorteio': resultados_sorteio,
        'horario_conclusao': request.session.get('horario_conclusao', '')
    }

    return render(request, 'buriti/buriti_sorteio.html', context)


def buriti_excel(request):
    # Caminho do modelo Excel
    caminho_modelo = 'setup/static/assets/modelos/sorteioburiti.xlsx'

    # Carregar o modelo existente
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Pegar todos os sorteios ordenados por ID do apartamento
    resultados_sorteio = list(Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all())

    # Pegar o horário de conclusão do sorteio
    horario_conclusao = request.session.get('horario_conclusao', 'Horário não disponível')
    ws['A8'] = f"Sorteio realizado em: {horario_conclusao}"

    # Limpar linhas antigas a partir da linha 10
    linha_inicial = 10
    linha_maxima = max(ws.max_row, linha_inicial + len(resultados_sorteio) + 50)
    
    for linha in range(linha_inicial, linha_maxima + 1):
        ws[f'A{linha}'] = None
        ws[f'B{linha}'] = None
        ws[f'C{linha}'] = None
        ws[f'D{linha}'] = None

    # Preencher dados
    linha = 10
    for sorteio in resultados_sorteio:
        ws[f'A{linha}'] = f"Bloco {sorteio.apartamento.bloco} - {sorteio.apartamento.numero}"
        ws[f'B{linha}'] = sorteio.vaga.numero
        ws[f'C{linha}'] = sorteio.vaga.bloco if sorteio.vaga.bloco else '-'
        ws[f'D{linha}'] = 'Buriti'  # Nome do condomínio
        linha += 1

    # Configurar a resposta para o download do Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sorteio_buriti.xlsx"'

    # Salvar o arquivo Excel na resposta
    wb.save(response)

    return response


def buriti_qrcode(request):
    # Obter o apartamento selecionado através do filtro (via GET)
    numero_apartamento = request.GET.get('apartamento')
    
    # Obter apartamentos que têm sorteio
    apartamentos_com_sorteio = Sorteio.objects.values_list('apartamento_id', flat=True).distinct()
    
    # Obter objetos Apartamento que têm sorteio (ordenados por ID)
    apartamentos_disponiveis = Apartamento.objects.filter(
        id__in=apartamentos_com_sorteio
    ).order_by('id')

    # Inicializar a variável de resultados filtrados como uma lista vazia
    resultados_filtrados = []

    # Se o apartamento foi selecionado, realizar a filtragem dos resultados do sorteio
    if numero_apartamento:
        # Buscar os sorteios para o apartamento selecionado com select_related para otimizar
        # O número do apartamento pode ser único ou pode precisar do bloco, mas vamos buscar pelo número
        resultados_filtrados = Sorteio.objects.filter(
            apartamento__numero=numero_apartamento
        ).select_related('apartamento', 'vaga').order_by('apartamento__id')

    return render(request, 'buriti/buriti_qrcode.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': numero_apartamento,
        'apartamentos_disponiveis': apartamentos_disponiveis,
    })


def buriti_zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        request.session['sorteio_iniciado'] = False
        request.session.pop('horario_conclusao', None)
        return redirect('buriti_sorteio')
    else:
        return render(request, 'buriti/buriti_zerar.html')

