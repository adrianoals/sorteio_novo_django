from django.shortcuts import render, redirect
from django.utils import timezone
import random
from django.http import HttpResponse
from django.urls import reverse
# from openpyxl import Workbook  # Para gerar o Excel
from openpyxl import load_workbook
from io import BytesIO  # Para manipular imagens em memória
from gran_vitta.models import Apartamento, Vaga, Sorteio


def gran_vitta_sorteio(request):
    if request.method == 'POST':
        # Marcar o sorteio como iniciado
        request.session['sorteio_iniciado'] = True

        # Redirecionar imediatamente para renderizar o estado do sorteio
        return redirect(reverse('gran_vitta_sorteio'))

    # Sorteio sendo iniciado
    sorteio_iniciado = request.session.get('sorteio_iniciado', False)

    if sorteio_iniciado:
        # Limpar registros anteriores de sorteio (OTIMIZADO)
        Sorteio.objects.all().delete()

        # Obtenha todos os apartamentos e vagas disponíveis
        apartamentos = list(Apartamento.objects.all())
        vagas_disponiveis = list(Vaga.objects.all())

        # Separar apartamentos e vagas PNE
        apartamentos_pne = [apt for apt in apartamentos if apt.is_pne]
        apartamentos_regulares = [apt for apt in apartamentos if not apt.is_pne]
        vagas_pne = [vaga for vaga in vagas_disponiveis if vaga.is_pne]
        vagas_regulares = [vaga for vaga in vagas_disponiveis if not vaga.is_pne]

        # Lista para armazenar todos os sorteios (bulk insert)
        sorteios_para_criar = []

        # Sortear vagas PNE entre os apartamentos PNE
        if apartamentos_pne and vagas_pne:
            # Embaralhar aleatoriamente os apartamentos PNE e as vagas PNE
            random.shuffle(apartamentos_pne)
            random.shuffle(vagas_pne)

            # Realizar sorteio para as vagas PNE
            for i, apartamento in enumerate(apartamentos_pne):
                if i < len(vagas_pne):
                    vaga_pne = vagas_pne[i]
                    sorteios_para_criar.append(
                        Sorteio(apartamento=apartamento, vaga=vaga_pne)
                    )
                else:
                    # Se não houver mais vagas PNE, adiciona às vagas regulares
                    apartamentos_regulares.append(apartamento)

            # Adicionar vagas PNE não usadas ao pool de vagas regulares
            vagas_pne_nao_usadas = vagas_pne[len(apartamentos_pne):]
            vagas_regulares.extend(vagas_pne_nao_usadas)
        elif vagas_pne:
            # Se não há apartamentos PNE, todas as vagas PNE vão para o pool geral
            vagas_regulares.extend(vagas_pne)

        # Sortear vagas regulares para os apartamentos restantes
        random.shuffle(vagas_regulares)
        for i, apartamento in enumerate(apartamentos_regulares):
            if i < len(vagas_regulares):
                vaga = vagas_regulares[i]
                sorteios_para_criar.append(
                    Sorteio(apartamento=apartamento, vaga=vaga)
                )

        # Criar todos os sorteios de uma vez (MUITO MAIS RÁPIDO)
        if sorteios_para_criar:
            Sorteio.objects.bulk_create(sorteios_para_criar)

        # Marcar o sorteio como concluído
        request.session['sorteio_iniciado'] = False
        request.session['horario_conclusao'] = timezone.localtime().strftime("%d/%m/%Y às %Hh %Mmin e %Ss")

    # Exibir os resultados do sorteio ordenados por ID do apartamento
    vagas_atribuidas = Sorteio.objects.exists()
    resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id') if vagas_atribuidas else None

    context = {
        'sorteio_iniciado': sorteio_iniciado,
        'vagas_atribuidas': vagas_atribuidas,
        'resultados_sorteio': resultados_sorteio,
        'horario_conclusao': request.session.get('horario_conclusao', '')  # Exibe o horário de conclusão
    }

    return render(request, 'gran_vitta/gran_vitta_sorteio.html', context)


def gran_vitta_excel(request):
    # Caminho do modelo Excel
    caminho_modelo = 'setup/static/assets/modelos/sorteiogranvitta.xlsx'

    # Carregar o modelo existente
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

    # Pegar o horário de conclusão do sorteio
    horario_conclusao = request.session.get('horario_conclusao', 'Horário não disponível')
    ws['A8'] = f"Sorteio realizado em: {horario_conclusao}"

    # Começar a partir da linha 10 (baseado no layout do seu modelo)
    linha = 10
    for sorteio in resultados_sorteio:
        ws[f'A{linha}'] = f'Unidade {sorteio.apartamento.numero}' # Número do apartamento
        ws[f'B{linha}'] = sorteio.vaga.numero  # Número da vaga
        ws[f'C{linha}'] = sorteio.vaga.subsolo # Subsolo
        linha += 1

    # Configurar a resposta para o download do Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sorteio_gran_vitta.xlsx"'

    # Salvar o arquivo Excel na resposta
    wb.save(response)

    return response



def gran_vitta_qrcode(request):
    # Obter todos os apartamentos para preencher o dropdown
    apartamentos_disponiveis = Apartamento.objects.all()
    
    # Obter o apartamento selecionado através do filtro (via GET)
    numero_apartamento = request.GET.get('apartamento')

    # Inicializar a variável de resultados filtrados como uma lista vazia
    resultados_filtrados = []

    # Se o apartamento foi selecionado, realizar a filtragem dos resultados do sorteio
    if numero_apartamento:
        # Buscar os sorteios para o apartamento selecionado
        resultados_filtrados = Sorteio.objects.filter(apartamento__numero=numero_apartamento)

    return render(request, 'gran_vitta/gran_vitta_qrcode.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': numero_apartamento,
        'apartamentos_disponiveis': apartamentos_disponiveis,
    })



def gran_vitta_zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        return redirect('gran_vitta_sorteio')
    else:
        return render(request, 'gran_vitta/gran_vitta_zerar.html')
    

