from django.shortcuts import render, redirect
from django.utils import timezone
import random
from django.http import HttpResponse
from django.urls import reverse
# from openpyxl import Workbook  # Para gerar o Excel
from openpyxl import load_workbook
from io import BytesIO  # Para manipular imagens em memória
from .models import Apartamento, Vaga, Sorteio

# View para realizar o sorteio
def sky_view_sorteio(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio
        Sorteio.objects.all().delete()

        # Carregar todos os apartamentos e vagas de uma vez (otimização)
        todos_apartamentos = list(Apartamento.objects.all())
        todas_vagas_simples = list(Vaga.objects.filter(tipo_vaga='Simples'))
        todas_vagas_duplas = list(Vaga.objects.filter(tipo_vaga='Dupla'))

        # Usar set para rastrear IDs de vagas e apartamentos já atribuídos (mais eficiente)
        vagas_atribuidas_ids = set()
        apartamentos_atribuidos_ids = set()
        sorteios_para_criar = []

        # ============================================
        # 1. ALOCAR APARTAMENTO 505 (vaga restrita)
        # ============================================
        apt_505 = None
        for apt in todos_apartamentos:
            if apt.numero == '505':
                apt_505 = apt
                break

        if apt_505:
            # Vagas possíveis para o 505: 9, 10 ou 12 do 1º Subsolo
            vagas_505 = []
            for v in todas_vagas_simples:
                if v.subsolo == '1º Subsolo' and v.id not in vagas_atribuidas_ids:
                    # Extrair número da vaga (pode ser "Vaga 9", "Vaga 09", "9", etc.)
                    numero_limpo = v.numero.replace('Vaga ', '').strip()
                    # Verificar se é uma das vagas permitidas (9, 10 ou 12)
                    if numero_limpo in ['9', '09', '10', '12']:
                        vagas_505.append(v)
            
            if vagas_505:
                vaga_505_encontrada = random.choice(vagas_505)
                sorteios_para_criar.append(Sorteio(apartamento=apt_505, vaga=vaga_505_encontrada))
                vagas_atribuidas_ids.add(vaga_505_encontrada.id)
                apartamentos_atribuidos_ids.add(apt_505.id)
                print(f"✅ Apartamento 505 → {vaga_505_encontrada.numero} ({vaga_505_encontrada.subsolo})")

        # ============================================
        # 2. APARTAMENTOS COM DIREITO A DUAS VAGAS LIVRES
        # ============================================
        apts_duas_vagas = [
            apt for apt in todos_apartamentos
            if apt.direito_duas_vagas_livres and apt.id not in apartamentos_atribuidos_ids
        ]
        random.shuffle(apts_duas_vagas)

        # Filtrar vagas disponíveis uma única vez (otimização)
        vagas_disponiveis_duas = [v for v in todas_vagas_simples if v.id not in vagas_atribuidas_ids]
        random.shuffle(vagas_disponiveis_duas)
        indice_vagas = 0

        for apt in apts_duas_vagas:
            # Atribuir 2 vagas simples
            if indice_vagas + 1 < len(vagas_disponiveis_duas):
                vaga1 = vagas_disponiveis_duas[indice_vagas]
                vaga2 = vagas_disponiveis_duas[indice_vagas + 1]
                
                sorteios_para_criar.append(Sorteio(apartamento=apt, vaga=vaga1))
                sorteios_para_criar.append(Sorteio(apartamento=apt, vaga=vaga2))
                vagas_atribuidas_ids.add(vaga1.id)
                vagas_atribuidas_ids.add(vaga2.id)
                apartamentos_atribuidos_ids.add(apt.id)
                indice_vagas += 2
                print(f"✅ Apartamento {apt.numero} → 2 vagas: {vaga1.numero} e {vaga2.numero}")

        # ============================================
        # 3. APARTAMENTOS COM DIREITO A VAGA DUPLA
        # ============================================
        apts_vaga_dupla = [
            apt for apt in todos_apartamentos
            if apt.direito_vaga_dupla and apt.id not in apartamentos_atribuidos_ids
        ]
        random.shuffle(apts_vaga_dupla)

        # Filtrar vagas duplas disponíveis uma única vez (otimização)
        vagas_duplas_disponiveis = [
            v for v in todas_vagas_duplas
            if v.id not in vagas_atribuidas_ids
        ]
        random.shuffle(vagas_duplas_disponiveis)
        indice_duplas = 0

        for apt in apts_vaga_dupla:
            if indice_duplas < len(vagas_duplas_disponiveis):
                vaga_dupla = vagas_duplas_disponiveis[indice_duplas]
                sorteios_para_criar.append(Sorteio(apartamento=apt, vaga=vaga_dupla))
                vagas_atribuidas_ids.add(vaga_dupla.id)
                apartamentos_atribuidos_ids.add(apt.id)
                indice_duplas += 1
                print(f"✅ Apartamento {apt.numero} → Vaga Dupla {vaga_dupla.numero}")

        # ============================================
        # 4. DEMAIS APARTAMENTOS (1 vaga simples cada)
        # ============================================
        apts_restantes = [
            apt for apt in todos_apartamentos
            if apt.id not in apartamentos_atribuidos_ids
        ]
        random.shuffle(apts_restantes)

        # Filtrar vagas simples disponíveis uma única vez (otimização)
        vagas_simples_disponiveis = [
            v for v in todas_vagas_simples
            if v.id not in vagas_atribuidas_ids
        ]
        random.shuffle(vagas_simples_disponiveis)
        indice_simples = 0

        for apt in apts_restantes:
            if indice_simples < len(vagas_simples_disponiveis):
                vaga_simples = vagas_simples_disponiveis[indice_simples]
                sorteios_para_criar.append(Sorteio(apartamento=apt, vaga=vaga_simples))
                apartamentos_atribuidos_ids.add(apt.id)
                indice_simples += 1
                print(f"✅ Apartamento {apt.numero} → {vaga_simples.numero} ({vaga_simples.subsolo})")

        # ============================================
        # 5. BULK CREATE (otimização de performance)
        # ============================================
        if sorteios_para_criar:
            Sorteio.objects.bulk_create(sorteios_para_criar)
            print(f"\n✅ Total de {len(sorteios_para_criar)} sorteios criados com sucesso!")

        # Marcar o sorteio como iniciado e armazenar o horário de conclusão
        request.session['sorteio_iniciado'] = True
        request.session['horario_conclusao'] = timezone.localtime().strftime("%d/%m/%Y às %Hh %Mmin e %Ss")

        # Redireciona para a mesma página após o sorteio
        return redirect(reverse('sky_view_sorteio'))

    # Se o método for GET, exibe os resultados ou a interface de sorteio
    sorteio_iniciado = request.session.get('sorteio_iniciado', False)
    vagas_atribuidas = Sorteio.objects.exists()
    # resultados_sorteio = Sorteio.objects.all() if vagas_atribuidas else None
    resultados_sorteio = Sorteio.objects.order_by('apartamento__id') if vagas_atribuidas else None

    context = {
        'sorteio_iniciado': sorteio_iniciado,
        'vagas_atribuidas': vagas_atribuidas,
        'resultados_sorteio': resultados_sorteio,
        'horario_conclusao': request.session.get('horario_conclusao', '')  # Exibe o horário de conclusão
    }

    return render(request, 'sky_view/sky_view_sorteio.html', context)



def sky_view_excel(request):
    # Caminho do modelo Excel
    caminho_modelo = 'setup/static/assets/modelos/sorteioskyview.xlsx'

    # Carregar o modelo existente
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Pegar todos os sorteios ordenados por apartamento
    resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__numero')

    # Pegar o horário de conclusão do sorteio
    horario_conclusao = request.session.get('horario_conclusao', 'Horário não disponível')
    ws['A8'] = f"Sorteio realizado em: {horario_conclusao}"

    # Limpar linhas antigas a partir da linha 10 para evitar dados antigos
    linha_inicial = 10
    linha_maxima = ws.max_row
    # Limpar todas as células das colunas A, B, C, D a partir da linha 10
    for linha in range(linha_inicial, linha_maxima + 1):
        ws[f'A{linha}'] = None
        ws[f'B{linha}'] = None
        ws[f'C{linha}'] = None
        ws[f'D{linha}'] = None

    # Começar a partir da linha 10 (baseado no layout do seu modelo)
    linha = 10
    for sorteio in resultados_sorteio:
        ws[f'A{linha}'] = sorteio.apartamento.numero  # Número do apartamento
        ws[f'B{linha}'] = sorteio.vaga.numero  # Número da vaga
        ws[f'C{linha}'] = sorteio.vaga.subsolo  # Subsolo
        ws[f'D{linha}'] = sorteio.vaga.tipo_vaga  # Tipo da vaga
        linha += 1

    # Configurar a resposta para o download do Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sorteio_sky_view.xlsx"'

    # Salvar o arquivo Excel na resposta
    wb.save(response)

    return response



def sky_view_qrcode(request):
    # Obter todos os apartamentos para preencher o dropdown
    apartamentos_disponiveis = Apartamento.objects.all()
    
    # Obter o apartamento selecionado através do filtro (via GET)
    numero_apartamento = request.GET.get('apartamento')

    # Inicializar a variável de resultados filtrados como uma lista vazia
    resultados_filtrados = []

    # Se o apartamento foi selecionado, realizar a filtragem dos resultados do sorteio
    if numero_apartamento:
        # Buscar os sorteios para o apartamento selecionado
        resultados_filtrados = Sorteio.objects.filter(apartamento__numero=numero_apartamento)

    return render(request, 'sky_view/sky_view_qrcode.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': numero_apartamento,
        'apartamentos_disponiveis': apartamentos_disponiveis,
    })



def sky_view_zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        return redirect('sky_view_sorteio')
    else:
        return render(request, 'sky_view/sky_view_zerar.html')
    

