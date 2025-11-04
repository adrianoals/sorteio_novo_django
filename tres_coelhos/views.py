from django.shortcuts import render, redirect
from django.utils import timezone
from django.contrib import messages
import random
from django.http import HttpResponse
# from openpyxl import Workbook  # Para gerar o Excel
from openpyxl import load_workbook
import qrcode  # Para gerar o QR Code
from io import BytesIO  # Para manipular imagens em memória
from .models import Apartamento, Vaga, Sorteio, SorteioDupla, DuplaApartamentos


def tres_coelhos_sorteio(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio
        Sorteio.objects.all().delete()
        print("Sorteios anteriores apagados.")

        # Obter todos os apartamentos e vagas LIVRES
        apartamentos_pne = list(Apartamento.objects.filter(is_pne=True))
        apartamentos_demais = list(Apartamento.objects.filter(is_pne=False))
        vagas_pne_livres = list(Vaga.objects.filter(especial='PNE', tipo='LIVRE'))
        vagas_livres_normais = list(Vaga.objects.filter(especial='NORMAL', tipo='LIVRE'))

        print(f"Apartamentos PNE: {len(apartamentos_pne)}")
        print(f"Apartamentos Demais: {len(apartamentos_demais)}")
        print(f"Vagas PNE Livres: {len(vagas_pne_livres)}")
        print(f"Vagas Livres Normais: {len(vagas_livres_normais)}")

        # SIMPLIFICAÇÃO: Mapeamento fixo de subsolo -> vaga PNE livre
        # Como existe exatamente 1 vaga PNE livre por subsolo, pré-mapeamos
        # Vagas PNE esperadas: Subsolo 1 → ID 43, Subsolo 2 → ID 95
        vaga_pne_esperada = {1: 43, 2: 95}  # Mapeamento esperado: subsolo -> id_vaga
        
        vaga_pne_por_subsolo = {}
        for vaga_pne in vagas_pne_livres:
            if vaga_pne.subsolo not in vaga_pne_por_subsolo:
                vaga_pne_por_subsolo[vaga_pne.subsolo] = vaga_pne
                # Validar se encontrou a vaga PNE esperada
                id_esperado = vaga_pne_esperada.get(vaga_pne.subsolo)
                status_validacao = "✅ CORRETO" if vaga_pne.id == id_esperado else f"⚠️ ESPERADO ID {id_esperado}"
                print(f"Mapeamento fixo: Subsolo {vaga_pne.subsolo} → Vaga PNE {vaga_pne.numero} (ID: {vaga_pne.id}) {status_validacao}")
            else:
                print(f"⚠️ ATENÇÃO: Múltiplas vagas PNE no Subsolo {vaga_pne.subsolo} - usando a primeira encontrada")

        # Lista para armazenar todos os sorteios (bulk insert - otimizado)
        sorteios_para_criar = []

        # REGRA 1: CRITÉRIO DE SUBSOLO (PRIORIDADE MÁXIMA)
        # Processar cada subsolo separadamente
        subsolos = set()
        subsolos.update([apt.subsolo for apt in apartamentos_pne + apartamentos_demais if apt.subsolo])
        subsolos.update([vaga.subsolo for vaga in vagas_livres_normais])
        subsolos.update(vaga_pne_por_subsolo.keys())

        print(f"Subsolos encontrados: {sorted(subsolos)}")

        for subsolo in sorted(subsolos):
            print(f"\n--- Processando Subsolo {subsolo} ---")

            # Separar apartamentos e vagas deste subsolo
            pne_subsolo = [apt for apt in apartamentos_pne if apt.subsolo == subsolo]
            demais_subsolo = [apt for apt in apartamentos_demais if apt.subsolo == subsolo]
            vagas_normais_subsolo = [vaga for vaga in vagas_livres_normais if vaga.subsolo == subsolo]

            print(f"Subsolo {subsolo}: {len(pne_subsolo)} PNE, {len(demais_subsolo)} Demais")
            print(f"Subsolo {subsolo}: {len(vagas_normais_subsolo)} vagas livres normais")

            # Inicializar PNE remanescentes (sempre inicializado para evitar erros)
            pne_remanescentes_subsolo = []

            # ETAPA 1: 1 PNE do subsolo ocupa a vaga PNE fixa daquele subsolo
            vaga_pne_fixa = vaga_pne_por_subsolo.get(subsolo)
            
            if vaga_pne_fixa and pne_subsolo:
                # Sortear 1 PNE para ocupar a vaga PNE fixa
                random.shuffle(pne_subsolo)
                pne_escolhido = pne_subsolo[0]
                
                # Atribuir vaga PNE fixa ao PNE escolhido
                sorteios_para_criar.append(
                    Sorteio(apartamento=pne_escolhido, vaga=vaga_pne_fixa)
                )
                print(f"Subsolo {subsolo}: Sorteado PNE {pne_escolhido.numero} → Vaga PNE {vaga_pne_fixa.numero} (fixa)")

                # PNE restantes do mesmo subsolo concorrem exclusivamente às vagas livres remanescentes
                pne_remanescentes_subsolo = pne_subsolo[1:]
                
                if pne_remanescentes_subsolo:
                    print(f"Subsolo {subsolo}: {len(pne_remanescentes_subsolo)} PNE remanescentes que devem receber vaga livre normal")
            elif pne_subsolo:
                # Há PNE mas não há vaga PNE fixa neste subsolo - todos os PNE vão para sorteio de vagas livres
                pne_remanescentes_subsolo = pne_subsolo
                print(f"Subsolo {subsolo}: {len(pne_remanescentes_subsolo)} PNE sem vaga PNE fixa - todos devem receber vaga livre normal")
            elif vaga_pne_fixa:
                # Há vaga PNE fixa mas não há PNE - vaga PNE é liberada para qualquer apartamento do subsolo
                print(f"Subsolo {subsolo}: Vaga PNE fixa disponível mas sem PNE - vaga será liberada para sorteio geral")
                # Adicionar vaga PNE ao pool de vagas livres do subsolo
                vagas_normais_subsolo.append(vaga_pne_fixa)

            # ETAPA 1b: PNE remanescentes recebem vagas livres (PRIORIDADE - antes dos demais)
            if pne_remanescentes_subsolo and vagas_normais_subsolo:
                random.shuffle(pne_remanescentes_subsolo)
                random.shuffle(vagas_normais_subsolo)
                
                # Atribuir vagas livres aos PNE remanescentes (número máximo é o menor entre vagas e PNE)
                num_atribuicoes_pne_remanescentes = min(len(pne_remanescentes_subsolo), len(vagas_normais_subsolo))
                
                for i in range(num_atribuicoes_pne_remanescentes):
                    # PNE remanescente recebe vaga livre do mesmo subsolo
                    sorteios_para_criar.append(
                        Sorteio(apartamento=pne_remanescentes_subsolo[i], vaga=vagas_normais_subsolo[i])
                    )
                    print(f"Subsolo {subsolo}: Sorteado PNE remanescente {pne_remanescentes_subsolo[i].numero} → Vaga Livre {vagas_normais_subsolo[i].numero}")
                
                # Remover vagas já atribuídas dos PNE remanescentes
                vagas_normais_subsolo = vagas_normais_subsolo[num_atribuicoes_pne_remanescentes:]
                
                if len(pne_remanescentes_subsolo) > num_atribuicoes_pne_remanescentes:
                    pne_sem_vaga = len(pne_remanescentes_subsolo) - num_atribuicoes_pne_remanescentes
                    print(f"Subsolo {subsolo}: ⚠️ {pne_sem_vaga} PNE remanescente(s) sem vaga livre (vagas esgotadas neste subsolo)")

            # ETAPA 1c: Demais apartamentos disputam as vagas livres restantes
            if demais_subsolo and vagas_normais_subsolo:
                random.shuffle(demais_subsolo)
                random.shuffle(vagas_normais_subsolo)
                
                # Atribuir vagas livres aos demais apartamentos (número máximo é o menor entre vagas e apartamentos)
                num_atribuicoes_demais = min(len(demais_subsolo), len(vagas_normais_subsolo))
                
                for i in range(num_atribuicoes_demais):
                    # Demais apartamento recebe vaga livre do mesmo subsolo
                    sorteios_para_criar.append(
                        Sorteio(apartamento=demais_subsolo[i], vaga=vagas_normais_subsolo[i])
                    )
                    print(f"Subsolo {subsolo}: Sorteado {demais_subsolo[i].numero} → Vaga Livre {vagas_normais_subsolo[i].numero}")
                
                if len(demais_subsolo) > num_atribuicoes_demais:
                    demais_sem_vaga = len(demais_subsolo) - num_atribuicoes_demais
                    print(f"Subsolo {subsolo}: {demais_sem_vaga} apartamento(s) sem vaga livre → vai(ão) para sorteio de duplas")

        # Criar todos os sorteios de uma vez (MUITO MAIS RÁPIDO)
        if sorteios_para_criar:
            Sorteio.objects.bulk_create(sorteios_para_criar)

        # Armazenar informações do sorteio na sessão
        request.session['sorteio_iniciado'] = True
        request.session['horario_conclusao'] = timezone.localtime().strftime("%d/%m/%Y às %Hh %Mmin e %Ss")

        return redirect('tres_coelhos_sorteio')

    else:
        sorteio_iniciado = request.session.get('sorteio_iniciado', False)
        resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()
        vagas_atribuidas = resultados_sorteio.exists()

        return render(request, 'tres_coelhos/tres_coelhos_sorteio.html', {
            'resultados_sorteio': resultados_sorteio,
            'vagas_atribuidas': vagas_atribuidas,
            'sorteio_iniciado': sorteio_iniciado,
            'horario_conclusao': request.session.get('horario_conclusao', '')
        })



def tres_coelhos_excel(request):
    # Caminho do modelo Excel
    caminho_modelo = 'setup/static/assets/modelos/sorteio_novo.xlsx'

    # Carregar o modelo existente
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

    # Pegar o horário de conclusão do sorteio
    horario_conclusao = request.session.get('horario_conclusao', 'Horário não disponível')
    ws['B8'] = f"Sorteio realizado em: {horario_conclusao}"

    # Começar a partir da linha 10 (baseado no layout do seu modelo)
    linha = 10
    for sorteio in resultados_sorteio:
        ws[f'B{linha}'] = sorteio.apartamento.numero  # Número do apartamento
        ws[f'C{linha}'] = sorteio.vaga.numero  # Número da vaga
        ws[f'D{linha}'] = f'Subsolo {sorteio.vaga.subsolo}'  # Subsolo
        ws[f'E{linha}'] = sorteio.vaga.get_tipo_display()  # Tipo da vaga
        ws[f'F{linha}'] = sorteio.vaga.get_especial_display()  # Especialidade da vaga
        linha += 1

    # Configurar a resposta para o download do Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="resultado_sorteio_tres_coelhos.xlsx"'

    # Salvar o arquivo Excel na resposta
    wb.save(response)

    return response



def tres_coelhos_qrcode(request):
    # Obter o apartamento selecionado através do filtro (via GET)
    numero_apartamento = request.GET.get('apartamento')
    
    # Obter apartamentos que têm sorteio em AMBAS as tabelas (Sorteio e SorteioDupla)
    apartamentos_com_sorteio = set()
    
    # Buscar apartamentos da tabela Sorteio (Fase 1 - vagas livres)
    apartamentos_fase1 = Sorteio.objects.values_list('apartamento_id', flat=True).distinct()
    apartamentos_com_sorteio.update(apartamentos_fase1)
    
    # Buscar apartamentos da tabela SorteioDupla (Fase 2 - vagas duplas)
    apartamentos_fase2 = SorteioDupla.objects.values_list('apartamento_id', flat=True).distinct()
    apartamentos_com_sorteio.update(apartamentos_fase2)
    
    # Obter objetos Apartamento que têm sorteio (ordenados por número)
    apartamentos_disponiveis = Apartamento.objects.filter(
        id__in=apartamentos_com_sorteio
    ).order_by('numero')
    
    # Ordenar por número convertendo para int para ordenação numérica correta (1, 2, ..., 812)
    apartamentos_disponiveis = sorted(
        apartamentos_disponiveis,
        key=lambda apt: int(apt.numero) if apt.numero.isdigit() else float('inf')
    )
    
    # Buscar TODOS os sorteios de TODAS as tabelas para encontrar duplas
    # Isso é necessário porque um apartamento pode fazer dupla com outro que não está nos resultados filtrados
    todos_sorteios_fase1 = list(Sorteio.objects.select_related('apartamento', 'vaga').all())
    todos_sorteios_fase2 = list(SorteioDupla.objects.select_related('apartamento', 'vaga', 'vaga__dupla_com').all())
    todos_sorteios = todos_sorteios_fase1 + todos_sorteios_fase2
    
    # Inicializar a variável de resultados filtrados como uma lista vazia
    resultados_filtrados = []

    # Se o apartamento foi selecionado, realizar a filtragem dos resultados do sorteio
    if numero_apartamento:
        # Buscar os sorteios da Fase 1 (Sorteio) para o apartamento selecionado
        sorteios_fase1 = Sorteio.objects.filter(
            apartamento__numero=numero_apartamento
        ).select_related('apartamento', 'vaga').all()
        
        # Buscar os sorteios da Fase 2 (SorteioDupla) para o apartamento selecionado
        sorteios_fase2 = SorteioDupla.objects.filter(
            apartamento__numero=numero_apartamento
        ).select_related('apartamento', 'vaga', 'vaga__dupla_com').all()
        
        # Combinar os resultados de ambas as tabelas
        resultados_filtrados = list(sorteios_fase1) + list(sorteios_fase2)
        
        # Ordenar por tipo de sorteio (Fase 1 primeiro, depois Fase 2) e por número da vaga
        resultados_filtrados.sort(key=lambda x: (
            0 if isinstance(x, Sorteio) else 1,  # Sorteio vem primeiro
            int(x.vaga.numero) if x.vaga.numero.isdigit() else float('inf')
        ))

    return render(request, 'tres_coelhos/tres_coelhos_qrcode.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': numero_apartamento,
        'apartamentos_disponiveis': apartamentos_disponiveis,
        'todos_sorteios': todos_sorteios,  # Todos os sorteios para encontrar duplas
    })



def tres_coelhos_dupla(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio de duplas
        SorteioDupla.objects.all().delete()
        print("Sorteios anteriores apagados (duplas).")

        # Obter apartamentos que já foram sorteados na Fase 1 (têm vaga livre)
        apartamentos_com_vaga_livre = set(Sorteio.objects.values_list('apartamento_id', flat=True))
        
        # Obter todas as duplas pré-cadastradas
        duplas_pre_cadastradas = list(DuplaApartamentos.objects.all())
        
        # REGRA 2: Filtrar duplas válidas (ambos não foram sorteados na Fase 1)
        duplas_validas = []
        for dupla in duplas_pre_cadastradas:
            apt1_sorteado = dupla.apartamento_1.id in apartamentos_com_vaga_livre
            apt2_sorteado = dupla.apartamento_2.id in apartamentos_com_vaga_livre if dupla.apartamento_2 else False
            
            # Dupla é válida se nenhum dos apartamentos foi sorteado na Fase 1
            if not apt1_sorteado and not apt2_sorteado and dupla.apartamento_2:
                # REGRA 3: Verificar se ambos pertencem ao mesmo subsolo
                if dupla.apartamento_1.subsolo == dupla.apartamento_2.subsolo:
                    duplas_validas.append(dupla)
                else:
                    print(f"Dupla desfeita: Apartamentos {dupla.apartamento_1.numero} e {dupla.apartamento_2.numero} em subsolos diferentes")
            else:
                print(f"Dupla desfeita: Um dos apartamentos já foi sorteado na Fase 1 ({dupla.apartamento_1.numero}, {dupla.apartamento_2.numero if dupla.apartamento_2 else 'N/A'})")

        print(f"Duplas válidas (pré-configuradas): {len(duplas_validas)}")

        # Obter apartamentos que NÃO foram sorteados na Fase 1
        apartamentos_sem_vaga = list(Apartamento.objects.exclude(id__in=apartamentos_com_vaga_livre))

        # Obter vagas duplas disponíveis (não sorteadas na Fase 1)
        # Como SorteioDupla já foi limpo no início, não precisamos filtrar por SorteioDupla
        # Uma vaga dupla não pode estar em Sorteio (Fase 1 - vagas livres)
        vagas_duplas_ids_nao_usadas_na_fase1 = set(Vaga.objects.filter(tipo='DUPLA').exclude(
            id__in=Sorteio.objects.values_list('vaga_id', flat=True)
        ).values_list('id', flat=True))
        
        vagas_duplas = list(Vaga.objects.filter(id__in=vagas_duplas_ids_nao_usadas_na_fase1).select_related('dupla_com'))

        print(f"Apartamentos sem vaga (da Fase 1): {len(apartamentos_sem_vaga)}")
        print(f"Vagas duplas disponíveis: {len(vagas_duplas)}")

        # Lista para armazenar todos os sorteios (bulk insert - otimizado)
        sorteios_dupla_para_criar = []

        # REGRA 1: CRITÉRIO DE SUBSOLO (PRIORIDADE MÁXIMA)
        # Processar cada subsolo separadamente
        subsolos = set()
        subsolos.update([apt.subsolo for apt in apartamentos_sem_vaga if apt.subsolo])
        subsolos.update([vaga.subsolo for vaga in vagas_duplas])

        print(f"Subsolos encontrados: {sorted(subsolos)}")

        for subsolo in sorted(subsolos):
            print(f"\n--- Processando Subsolo {subsolo} (Vagas Duplas) ---")

            # Separar duplas válidas deste subsolo
            duplas_subsolo = [d for d in duplas_validas if d.apartamento_1.subsolo == subsolo]
            apartamentos_subsolo = [apt for apt in apartamentos_sem_vaga if apt.subsolo == subsolo]
            vagas_duplas_subsolo = [vaga for vaga in vagas_duplas if vaga.subsolo == subsolo]

            print(f"Subsolo {subsolo}: {len(duplas_subsolo)} duplas válidas, {len(apartamentos_subsolo)} apartamentos sem vaga")
            print(f"Subsolo {subsolo}: {len(vagas_duplas_subsolo)} vagas duplas disponíveis")
            print(f"Subsolo {subsolo}: Apartamentos IDs: {[apt.numero for apt in apartamentos_subsolo]}")

            # Criar lista de IDs de apartamentos já sorteados nesta etapa
            apartamentos_sorteados_dupla = set()
            
            # Criar dicionário para mapear vagas e seus pares
            vagas_disponiveis_set = set(vagas_duplas_subsolo)
            vagas_usadas = set()
            
            # REGRA 3: Sorteio das duplas pré-configuradas (primeiro)
            if duplas_subsolo:
                random.shuffle(duplas_subsolo)
                random.shuffle(vagas_duplas_subsolo)
                print(f"Subsolo {subsolo}: Processando {len(duplas_subsolo)} dupla(s) pré-configurada(s)")
                
                for dupla in duplas_subsolo:
                    # Encontrar par de vagas duplas disponíveis
                    vaga_encontrada = None
                    vaga_par = None
                    
                    # Buscar par de vagas duplas (uma deve ter dupla_com apontando para a outra)
                    for vaga in vagas_duplas_subsolo:
                        if vaga.id in vagas_usadas:
                            continue
                        
                        if vaga.dupla_com and vaga.dupla_com.id in [v.id for v in vagas_duplas_subsolo if v.id not in vagas_usadas]:
                            vaga_encontrada = vaga
                            vaga_par = vaga.dupla_com
                            break
                    
                    if vaga_encontrada and vaga_par:
                        # Criar sorteios da dupla
                        sorteios_dupla_para_criar.append(
                            SorteioDupla(apartamento=dupla.apartamento_1, vaga=vaga_encontrada)
                        )
                        sorteios_dupla_para_criar.append(
                            SorteioDupla(apartamento=dupla.apartamento_2, vaga=vaga_par)
                        )
                        
                        apartamentos_sorteados_dupla.add(dupla.apartamento_1.id)
                        apartamentos_sorteados_dupla.add(dupla.apartamento_2.id)
                        
                        # Marcar vagas como usadas (evita usar remove())
                        vagas_usadas.add(vaga_encontrada.id)
                        vagas_usadas.add(vaga_par.id)
                        
                        print(f"Subsolo {subsolo}: Dupla sorteada - {dupla.apartamento_1.numero} → Vaga {vaga_encontrada.numero}, {dupla.apartamento_2.numero} → Vaga {vaga_par.numero}")
                    else:
                        print(f"Subsolo {subsolo}: Dupla ({dupla.apartamento_1.numero}, {dupla.apartamento_2.numero}) sem par de vagas duplas disponível")
            else:
                print(f"Subsolo {subsolo}: Nenhuma dupla pré-configurada - todos os apartamentos serão atribuídos individualmente")

            # REGRA 4: Sorteio dos apartamentos restantes - Sem formar novas duplas (atribuição individual)
            apartamentos_restantes_subsolo = [apt for apt in apartamentos_subsolo if apt.id not in apartamentos_sorteados_dupla]
            
            print(f"Subsolo {subsolo}: Apartamentos restantes após duplas pré-configuradas: {len(apartamentos_restantes_subsolo)}")
            print(f"Subsolo {subsolo}: Apartamentos restantes IDs: {[apt.numero for apt in apartamentos_restantes_subsolo]}")
            print(f"Subsolo {subsolo}: Vagas usadas até agora: {len(vagas_usadas)}")
            
            random.shuffle(apartamentos_restantes_subsolo)
            
            # REGRA 4: Sorteio dos apartamentos restantes - Sem formar novas duplas (atribuição individual)
            # IMPORTANTE: Quando atribuímos uma vaga dupla individualmente, ela é atribuída sozinha.
            # A vaga par (dupla_com) permanece disponível para ser atribuída a outro apartamento.
            # Isso significa que se temos um par de vagas (Vaga A <-> Vaga B), ambas podem ser atribuídas individualmente.
            for apartamento in apartamentos_restantes_subsolo:
                # Filtrar vagas não usadas (atualizado a cada iteração)
                vagas_disponiveis = [v for v in vagas_duplas_subsolo if v.id not in vagas_usadas]
                
                if vagas_disponiveis:
                    # Apartamento recebe vaga dupla individualmente (sem formar par com outro apartamento)
                    # A vaga par (dupla_com) permanece disponível para outro apartamento
                    vaga_escolhida = vagas_disponiveis[0]  # Pega a primeira vaga disponível
                    sorteios_dupla_para_criar.append(
                        SorteioDupla(apartamento=apartamento, vaga=vaga_escolhida)
                    )
                    
                    apartamentos_sorteados_dupla.add(apartamento.id)
                    vagas_usadas.add(vaga_escolhida.id)
                    # NÃO marcamos a vaga par (dupla_com) como usada aqui, pois ela pode ser atribuída a outro apartamento
                    
                    par_info = f" (Par: {vaga_escolhida.dupla_com.numero})" if vaga_escolhida.dupla_com else ""
                    print(f"Subsolo {subsolo}: Apartamento {apartamento.numero} → Vaga Dupla {vaga_escolhida.numero}{par_info}")
                else:
                    print(f"Subsolo {subsolo}: Apartamento {apartamento.numero} sem vaga dupla disponível neste subsolo")
                    break  # Não há mais vagas disponíveis, parar o loop
            else:
                print(f"Subsolo {subsolo}: Nenhum apartamento restante para atribuir (todos já foram sorteados nas duplas pré-configuradas)")

        # Criar todos os sorteios de uma vez (MUITO MAIS RÁPIDO)
        if sorteios_dupla_para_criar:
            SorteioDupla.objects.bulk_create(sorteios_dupla_para_criar)
            print(f"\nTotal de sorteios de duplas criados: {len(sorteios_dupla_para_criar)}")
        else:
            print("\n⚠️ ATENÇÃO: Nenhum sorteio de dupla foi criado!")
        
        # Verificação final: contar vagas duplas não atribuídas
        total_vagas_duplas = Vaga.objects.filter(tipo='DUPLA').count()
        vagas_duplas_usadas_fase1 = Sorteio.objects.filter(vaga__tipo='DUPLA').count()
        vagas_duplas_usadas_fase2 = SorteioDupla.objects.count()
        
        print(f"\n📊 RESUMO FINAL:")
        print(f"Total de vagas duplas no sistema: {total_vagas_duplas}")
        print(f"Vagas duplas usadas na Fase 1 (erro): {vagas_duplas_usadas_fase1}")
        print(f"Vagas duplas atribuídas na Fase 2: {vagas_duplas_usadas_fase2}")
        print(f"Vagas duplas ainda disponíveis: {total_vagas_duplas - vagas_duplas_usadas_fase1 - vagas_duplas_usadas_fase2}")

        # Armazenar informações do sorteio na sessão
        request.session['sorteio_dupla_iniciado'] = True
        request.session['horario_conclusao_dupla'] = timezone.localtime().strftime("%d/%m/%Y às %Hh %Mmin e %Ss")

        return redirect('tres_coelhos_dupla')

    else:
        sorteio_iniciado = request.session.get('sorteio_dupla_iniciado', False)
        resultados_sorteio = SorteioDupla.objects.select_related('apartamento', 'vaga', 'vaga__dupla_com').order_by('apartamento__id').all()
        vagas_atribuidas = resultados_sorteio.exists()

        # Criar mapeamento de vaga_id -> apartamento_numero para facilitar busca no template
        vaga_para_apartamento = {}
        for sorteio in resultados_sorteio:
            vaga_para_apartamento[sorteio.vaga.id] = sorteio.apartamento.numero

        return render(request, 'tres_coelhos/tres_coelhos_dupla.html', {
            'resultados_sorteio': resultados_sorteio,
            'vagas_atribuidas': vagas_atribuidas,
            'sorteio_iniciado': sorteio_iniciado,
            'horario_conclusao': request.session.get('horario_conclusao_dupla', ''),
            'vaga_para_apartamento': vaga_para_apartamento
        })



def tres_coelhos_dupla_excel(request):
    # Caminho do modelo Excel
    caminho_modelo = 'setup/static/assets/modelos/sorteio_novo2.xlsx'

    # Carregar o modelo existente
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio = SorteioDupla.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

    # Pegar o horário de conclusão do sorteio
    horario_conclusao = request.session.get('horario_conclusao', 'Horário não disponível')
    ws['A8'] = f"Sorteio realizado em: {horario_conclusao}"

    # Começar a partir da linha 10 (baseado no layout do seu modelo)
    linha = 10
    for sorteio in resultados_sorteio:
        ws[f'A{linha}'] = sorteio.apartamento.numero  # Número do apartamento
        ws[f'B{linha}'] = sorteio.vaga.numero  # Número da vaga
        ws[f'C{linha}'] = f'Subsolo {sorteio.vaga.subsolo}'  # Subsolo
        ws[f'D{linha}'] = sorteio.vaga.get_tipo_display()  # Tipo da vaga
        ws[f'E{linha}'] = sorteio.vaga.get_especial_display()  # Especialidade da vaga

        # Adicionar "Dupla Com" (número da vaga associada, se houver)
        ws[f'F{linha}'] = sorteio.vaga.dupla_com.numero if sorteio.vaga.dupla_com else "N/A"  # Número da vaga dupla

        linha += 1

    # Configurar a resposta para o download do Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="resultado_sorteio_dupla_tres_coelhos.xlsx"'

    # Salvar o arquivo Excel na resposta
    wb.save(response)

    return response


def tres_coelhos_zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        SorteioDupla.objects.all().delete()
        return redirect('tres_coelhos_sorteio')
    else:
        return render(request, 'tres_coelhos/tres_coelhos_zerar.html')
    

def tres_coelhos_resultado(request):
    # Obtenha os resultados do sorteio e do sorteio duplo
    resultados_sorteio = list(Sorteio.objects.all())
    resultados_sorteio_dupla = list(SorteioDupla.objects.all())

    # Combine os resultados em uma única lista
    resultados_combinados = resultados_sorteio + resultados_sorteio_dupla

    # Ordene os resultados combinados pelo ID
    resultados_combinados.sort(key=lambda x: x.id)

    # Crie uma lista unificada com dados consistentes
    resultados_formatados = []
    for sorteio in resultados_combinados:
        apartamento_numero = sorteio.apartamento.numero if hasattr(sorteio, 'apartamento') else '-'
        vaga_numero = sorteio.vaga.numero if hasattr(sorteio, 'vaga') else '-'
        vaga_subsolo = sorteio.vaga.subsolo if hasattr(sorteio, 'vaga') else '-'
        tipo_vaga = sorteio.vaga.get_tipo_display() if hasattr(sorteio, 'vaga') else '-'
        especialidade = sorteio.vaga.get_especial_display() if hasattr(sorteio, 'vaga') else '-'
        dupla_com_numero = sorteio.vaga.dupla_com.numero if hasattr(sorteio.vaga, 'dupla_com') and sorteio.vaga.dupla_com else '-'

        resultados_formatados.append({
            'apartamento': apartamento_numero,
            'vaga': vaga_numero,
            'subsolo': vaga_subsolo,
            'tipo_vaga': tipo_vaga,
            'especialidade': especialidade,
            'dupla_com': dupla_com_numero
        })

    # Passe os dados formatados para o contexto
    context = {
        'resultados_combinados': resultados_formatados
    }

    return render(request, 'tres_coelhos/tres_coelhos_resultado.html', context)


def tres_coelhos_configurar_pne(request):
    """View para configurar quais apartamentos são PNE"""
    if request.method == 'POST':
        # Obter IDs dos apartamentos marcados como PNE
        apartamentos_pne_ids = request.POST.getlist('apartamentos_pne')
        
        # Converter para inteiros
        apartamentos_pne_ids = [int(id) for id in apartamentos_pne_ids if id.isdigit()]
        
        # Atualizar todos os apartamentos
        # Primeiro, desmarcar todos como PNE
        Apartamento.objects.all().update(is_pne=False)
        
        # Depois, marcar apenas os selecionados
        if apartamentos_pne_ids:
            Apartamento.objects.filter(id__in=apartamentos_pne_ids).update(is_pne=True)
        
        return redirect('tres_coelhos_configurar_pne')
    
    # Obter todos os apartamentos ordenados por número
    apartamentos = Apartamento.objects.all().order_by('numero')
    
    # Separar por subsolo
    apartamentos_subsolo_1 = [apt for apt in apartamentos if apt.subsolo == 1]
    apartamentos_subsolo_2 = [apt for apt in apartamentos if apt.subsolo == 2]
    apartamentos_sem_subsolo = [apt for apt in apartamentos if apt.subsolo is None]
    
    # Ordenar numericamente
    def ordenar_por_numero(apt):
        try:
            return int(apt.numero)
        except ValueError:
            return float('inf')
    
    apartamentos_subsolo_1 = sorted(apartamentos_subsolo_1, key=ordenar_por_numero)
    apartamentos_subsolo_2 = sorted(apartamentos_subsolo_2, key=ordenar_por_numero)
    apartamentos_sem_subsolo = sorted(apartamentos_sem_subsolo, key=ordenar_por_numero)
    
    return render(request, 'tres_coelhos/tres_coelhos_configurar_pne.html', {
        'apartamentos_subsolo_1': apartamentos_subsolo_1,
        'apartamentos_subsolo_2': apartamentos_subsolo_2,
        'apartamentos_sem_subsolo': apartamentos_sem_subsolo,
    })


def tres_coelhos_configurar_duplas(request):
    """View para configurar duplas de apartamentos por subsolo"""
    # Processar exclusão de dupla (via GET ou POST)
    dupla_id_excluir = request.GET.get('excluir_dupla') or request.POST.get('excluir_dupla')
    if dupla_id_excluir and dupla_id_excluir.isdigit():
        try:
            dupla = DuplaApartamentos.objects.get(id=int(dupla_id_excluir))
            dupla.delete()
            return redirect('tres_coelhos_configurar_duplas')
        except DuplaApartamentos.DoesNotExist:
            pass
    
    if request.method == 'POST':
        print("=" * 50)
        print("DEBUG: POST recebido!")
        print(f"DEBUG: Todos os campos POST: {list(request.POST.keys())}")
        print(f"DEBUG: Total de campos POST: {len(request.POST)}")
        print(f"DEBUG: csrfmiddlewaretoken presente: {'csrfmiddlewaretoken' in request.POST}")
        
        # Log detalhado de TODOS os valores POST
        print("\n=== TODOS OS VALORES POST ===")
        for key in request.POST.keys():
            value = request.POST.get(key)
            print(f"  {key} = '{value}' (tipo: {type(value).__name__})")
        print("=" * 50)
        
        try:
            # Obter dados do formulário
            duplas_data = []
            
            # Processar todas as chaves do POST que começam com 'dupla_'
            duplas_keys = {k for k in request.POST.keys() if k.startswith('dupla_')}
            
            print(f"\nDEBUG: Chaves encontradas (com 'dupla_'): {duplas_keys}")
            print(f"DEBUG: Total de chaves 'dupla_': {len(duplas_keys)}")
            
            # Debug: mostrar todos os valores
            for key in sorted(duplas_keys):
                value = request.POST.get(key, '')
                value_str = str(value).strip() if value else ''
                print(f"DEBUG: {key} = '{value_str}' (vazio: {not value_str})")
            
            # Agrupar por índice de dupla
            # Novo formato: dupla_sub1_0_apt1 ou dupla_sub2_1_apt2
            # parts = ['dupla', 'sub1', '0', 'apt1']
            duplas_dict = {}
            for key in duplas_keys:
                parts = key.split('_')
                if len(parts) == 4 and parts[0] == 'dupla' and parts[1].startswith('sub') and parts[3] in ['apt1', 'apt2']:
                    subsolo_str = parts[1]  # 'sub1' ou 'sub2'
                    index = parts[2]  # '0', '1', '2', etc.
                    campo = parts[3]  # 'apt1' ou 'apt2'
                    
                    # Criar chave única combinando subsolo e índice
                    key_unico = f"{subsolo_str}_{index}"  # 'sub1_0', 'sub2_1', etc.
                    
                    value = request.POST.get(key, '').strip()
                    if value:  # Só adiciona se houver valor
                        if key_unico not in duplas_dict:
                            duplas_dict[key_unico] = {}
                        duplas_dict[key_unico][campo] = value
                        print(f"DEBUG: Agrupado {key} -> duplas_dict[{key_unico}][{campo}] = {value}")
            
            print(f"DEBUG: Duplas agrupadas: {duplas_dict}")
            
            # Processar duplas coletadas
            for index, dupla in duplas_dict.items():
                apt1_id = str(dupla.get('apt1', '')).strip()
                apt2_id = str(dupla.get('apt2', '')).strip()
                
                print(f"DEBUG: Processando dupla {index}: apt1={apt1_id}, apt2={apt2_id}")
                
                # Validar que ambos os IDs existem e são numéricos
                if apt1_id and apt1_id.isdigit() and apt2_id and apt2_id.isdigit():
                    duplas_data.append((int(apt1_id), int(apt2_id)))
                    print(f"DEBUG: ✅ Dupla {index} válida: apt1={apt1_id}, apt2={apt2_id}")
                else:
                    print(f"DEBUG: ❌ Dupla {index} inválida - apt1={apt1_id}, apt2={apt2_id}")
            
            print(f"DEBUG: Duplas válidas encontradas: {len(duplas_data)}")
            
            if not duplas_data:
                messages.warning(request, 'Nenhuma dupla foi selecionada. Por favor, selecione pelo menos um apartamento em cada campo para formar uma dupla.')
                return redirect('tres_coelhos_configurar_duplas')
            
            # NÃO limpar duplas existentes - apenas adicionar novas
            # As duplas existentes são preservadas
            
            # Criar novas duplas (apenas se não existirem)
            duplas_para_criar = []
            apartamentos_em_duplas = set()
            
            # Coletar apartamentos que já estão em duplas
            duplas_existentes = DuplaApartamentos.objects.all()
            for dupla in duplas_existentes:
                apartamentos_em_duplas.add(dupla.apartamento_1.id)
                if dupla.apartamento_2:
                    apartamentos_em_duplas.add(dupla.apartamento_2.id)
            
            duplas_rejeitadas = []
            for apt1_id, apt2_id in duplas_data:
                # Verificar se algum dos apartamentos já está em uma dupla
                if apt1_id in apartamentos_em_duplas or apt2_id in apartamentos_em_duplas:
                    print(f"DEBUG: Pulando dupla - apartamento já está em dupla: {apt1_id}, {apt2_id}")
                    duplas_rejeitadas.append((apt1_id, apt2_id, "apartamento já está em dupla"))
                    continue
                    
                try:
                    apt1 = Apartamento.objects.get(id=apt1_id)
                    apt2 = Apartamento.objects.get(id=apt2_id)
                    
                    if apt1.id != apt2.id:
                        # Verificar se ambos têm o mesmo subsolo
                        if apt1.subsolo and apt2.subsolo and apt1.subsolo == apt2.subsolo:
                            duplas_para_criar.append(
                                DuplaApartamentos(apartamento_1=apt1, apartamento_2=apt2)
                            )
                            print(f"DEBUG: Dupla adicionada: {apt1.numero} ↔ {apt2.numero}")
                        else:
                            print(f"DEBUG: Dupla rejeitada - subsolos diferentes: {apt1.subsolo} vs {apt2.subsolo}")
                            duplas_rejeitadas.append((apt1_id, apt2_id, "subsolos diferentes"))
                    else:
                        print(f"DEBUG: Dupla rejeitada - mesmo apartamento: {apt1_id}")
                        duplas_rejeitadas.append((apt1_id, apt2_id, "mesmo apartamento"))
                except Apartamento.DoesNotExist as e:
                    print(f"DEBUG: Apartamento não encontrado: {apt1_id} ou {apt2_id}")
                    messages.error(request, f'Erro ao processar dupla: apartamento não encontrado.')
                    continue
            
            print(f"DEBUG: Total de duplas para criar: {len(duplas_para_criar)}")
            
            if duplas_para_criar:
                try:
                    # Criar as duplas uma por uma para debug melhor e validação
                    duplas_criadas_com_sucesso = 0
                    duplas_erros = []
                    
                    for dupla_obj in duplas_para_criar:
                        print(f"DEBUG: Criando dupla: {dupla_obj.apartamento_1.numero} ↔ {dupla_obj.apartamento_2.numero}")
                        print(f"DEBUG:   - Apt1 ID: {dupla_obj.apartamento_1.id}, Subsolo: {dupla_obj.apartamento_1.subsolo}")
                        print(f"DEBUG:   - Apt2 ID: {dupla_obj.apartamento_2.id}, Subsolo: {dupla_obj.apartamento_2.subsolo}")
                        
                        try:
                            # Verificar se a dupla já existe (evitar duplicatas)
                            dupla_existe = DuplaApartamentos.objects.filter(
                                apartamento_1=dupla_obj.apartamento_1,
                                apartamento_2=dupla_obj.apartamento_2
                            ).exists()
                            
                            if not dupla_existe:
                                dupla_obj.save()
                                duplas_criadas_com_sucesso += 1
                                print(f"DEBUG: ✅ Dupla criada com sucesso! ID: {dupla_obj.id}")
                            else:
                                print(f"DEBUG: ⚠️ Dupla já existe no banco - ignorando")
                                duplas_erros.append(f"Dupla {dupla_obj.apartamento_1.numero} ↔ {dupla_obj.apartamento_2.numero} já existe")
                        except Exception as save_error:
                            error_msg = str(save_error)
                            print(f"DEBUG: ❌ Erro ao criar dupla individual: {error_msg}")
                            import traceback
                            traceback.print_exc()
                            duplas_erros.append(f"Erro ao criar {dupla_obj.apartamento_1.numero} ↔ {dupla_obj.apartamento_2.numero}: {error_msg}")
                    
                    # Verificar se realmente foram criadas
                    total_depois = DuplaApartamentos.objects.count()
                    print(f"DEBUG: Total de duplas no banco após criar: {total_depois}")
                    print(f"DEBUG: Duplas criadas com sucesso: {duplas_criadas_com_sucesso}")
                    
                    if duplas_criadas_com_sucesso > 0:
                        messages.success(request, f'{duplas_criadas_com_sucesso} dupla(s) criada(s) com sucesso!')
                    if duplas_rejeitadas:
                        messages.warning(request, f'{len(duplas_rejeitadas)} dupla(s) foram rejeitada(s) (apartamentos já em duplas ou critérios não atendidos).')
                    if duplas_erros:
                        messages.warning(request, f'{len(duplas_erros)} dupla(s) não puderam ser criadas. Detalhes nos logs.')
                except Exception as db_error:
                    print(f"DEBUG: ❌ ERRO ao criar duplas no banco: {str(db_error)}")
                    import traceback
                    traceback.print_exc()
                    messages.error(request, f'Erro ao salvar no banco: {str(db_error)}')
            else:
                print(f"DEBUG: ⚠️ Nenhuma dupla para criar!")
                print(f"DEBUG: Duplas rejeitadas: {len(duplas_rejeitadas)}")
                for rej in duplas_rejeitadas:
                    print(f"DEBUG:   - {rej}")
                messages.warning(request, 'Nenhuma nova dupla foi criada. Verifique se os apartamentos selecionados estão corretos e não estão já em duplas.')
            
            print("=" * 50)
            
            return redirect('tres_coelhos_configurar_duplas')
            
        except Exception as e:
            print(f"ERRO ao processar formulário: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Erro ao processar formulário: {str(e)}')
            return redirect('tres_coelhos_configurar_duplas')
    
    # Obter duplas existentes organizadas por subsolo (ANTES de filtrar apartamentos)
    duplas_existentes = DuplaApartamentos.objects.select_related(
        'apartamento_1', 'apartamento_2'
    ).all()
    
    # Filtrar duplas por subsolo
    # Considerar duplas onde apartamento_1 tem o subsolo correto
    # E apartamento_2 existe e tem o mesmo subsolo
    duplas_subsolo_1 = []
    duplas_subsolo_2 = []
    
    for d in duplas_existentes:
        # Forçar avaliação dos objetos relacionados (evitar lazy loading)
        if hasattr(d, 'apartamento_1') and d.apartamento_1:
            # Forçar acesso ao subsolo para carregar o objeto
            subsolo_apt1 = d.apartamento_1.subsolo
            
            if subsolo_apt1 == 1:
                # Verificar se apartamento_2 existe e tem subsolo 1
                if hasattr(d, 'apartamento_2') and d.apartamento_2:
                    subsolo_apt2 = d.apartamento_2.subsolo
                    if subsolo_apt2 == 1:
                        duplas_subsolo_1.append(d)
                # Também incluir duplas incompletas do subsolo 1 para debug
                else:
                    print(f"DEBUG: Dupla subsolo 1 incompleta - ID: {d.id}, Apt1: {d.apartamento_1.numero}")
                    
            elif subsolo_apt1 == 2:
                # Verificar se apartamento_2 existe e tem subsolo 2
                if hasattr(d, 'apartamento_2') and d.apartamento_2:
                    subsolo_apt2 = d.apartamento_2.subsolo
                    if subsolo_apt2 == 2:
                        duplas_subsolo_2.append(d)
                    else:
                        print(f"DEBUG: Dupla subsolo 2 rejeitada - subsolos diferentes - ID: {d.id}, Apt1: {d.apartamento_1.numero} (subsolo {subsolo_apt1}), Apt2: {d.apartamento_2.numero if d.apartamento_2 else 'None'} (subsolo {subsolo_apt2})")
                else:
                    print(f"DEBUG: Dupla subsolo 2 incompleta - ID: {d.id}, Apt1: {d.apartamento_1.numero}")
    
    # Debug: logar duplas encontradas
    print(f"DEBUG: Total de duplas existentes: {duplas_existentes.count()}")
    print(f"DEBUG: Duplas subsolo 1: {len(duplas_subsolo_1)}")
    print(f"DEBUG: Duplas subsolo 2: {len(duplas_subsolo_2)}")
    for d in duplas_subsolo_2:
        apt1_num = d.apartamento_1.numero if d.apartamento_1 else 'None'
        apt2_num = d.apartamento_2.numero if d.apartamento_2 else 'None'
        print(f"DEBUG: Dupla subsolo 2 - ID: {d.id}, Apt1: {apt1_num} (subsolo: {d.apartamento_1.subsolo if d.apartamento_1 else None}), Apt2: {apt2_num} (subsolo: {d.apartamento_2.subsolo if d.apartamento_2 else None})")
    
    # Obter IDs dos apartamentos que já estão em duplas (para excluir dos dropdowns)
    apartamentos_em_duplas_ids = set()
    for dupla in duplas_existentes:
        apartamentos_em_duplas_ids.add(dupla.apartamento_1.id)
        if dupla.apartamento_2:
            apartamentos_em_duplas_ids.add(dupla.apartamento_2.id)
    
    # Obter todos os apartamentos ordenados por número e subsolo, EXCLUINDO os que já estão em duplas
    apartamentos_subsolo_1 = Apartamento.objects.filter(
        subsolo=1
    ).exclude(id__in=apartamentos_em_duplas_ids).order_by('numero')
    
    apartamentos_subsolo_2 = Apartamento.objects.filter(
        subsolo=2
    ).exclude(id__in=apartamentos_em_duplas_ids).order_by('numero')
    
    # Ordenar numericamente
    def ordenar_por_numero(apt):
        try:
            return int(apt.numero)
        except ValueError:
            return float('inf')
    
    apartamentos_subsolo_1 = sorted(list(apartamentos_subsolo_1), key=ordenar_por_numero)
    apartamentos_subsolo_2 = sorted(list(apartamentos_subsolo_2), key=ordenar_por_numero)
    
    return render(request, 'tres_coelhos/tres_coelhos_configurar_duplas.html', {
        'apartamentos_subsolo_1': apartamentos_subsolo_1,
        'apartamentos_subsolo_2': apartamentos_subsolo_2,
        'duplas_subsolo_1': duplas_subsolo_1,
        'duplas_subsolo_2': duplas_subsolo_2,
    })
