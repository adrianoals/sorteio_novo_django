from django.shortcuts import render, redirect
from django.utils import timezone
import random
from django.http import HttpResponse
from django.urls import reverse
from openpyxl import load_workbook
from io import BytesIO
from .models import Apartamento, Vaga, Sorteio
import time
import logging

logger = logging.getLogger(__name__)

# Apartamentos com vagas fixas (travadas)
APARTAMENTOS_VAGAS_FIXAS = {
    279: 413,  # Apartamento 23 -> Vaga 131
    283: 561,  # Apartamento 31 -> Vaga 279
    286: 564,  # Apartamento 34 -> Vaga 282
    289: 559,  # Apartamento 41 -> Vaga 277
    293: 538,  # Apartamento 45 -> Vaga 256
    296: 329,  # Apartamento 52 -> Vaga 47
    297: 537,  # Apartamento 53 -> Vaga 255
    314: 422,  # Apartamento 82 -> Vaga 140
    317: 511,  # Apartamento 85 -> Vaga 229
    318: 513,  # Apartamento 86 -> Vaga 231
    320: 522,  # Apartamento 92 -> Vaga 240
    321: 517,  # Apartamento 93 -> Vaga 235
    326: 515,  # Apartamento 102 -> Vaga 233
    330: 423,  # Apartamento 106 -> Vaga 141
    331: 364,  # Apartamento 111 -> Vaga 82
    333: 345,  # Apartamento 113 -> Vaga 63
    336: 292,  # Apartamento 116 -> Vaga 10
    342: 510,  # Apartamento 126 -> Vaga 228
    345: 428,  # Apartamento 133 -> Vaga 146
    347: 448,  # Apartamento 135 -> Vaga 166
    348: 516,  # Apartamento 136 -> Vaga 234
    349: 421,  # Apartamento 141 -> Vaga 139
    353: 521,  # Apartamento 145 -> Vaga 239
    355: 514,  # Apartamento 151 -> Vaga 232
    360: 435,  # Apartamento 156 -> Vaga 153
    362: 519,  # Apartamento 162 -> Vaga 237
    365: 417,  # Apartamento 165 -> Vaga 135
    366: 414,  # Apartamento 166 -> Vaga 132
    368: 518,  # Apartamento 172 -> Vaga 236
    369: 429,  # Apartamento 173 -> Vaga 147
    371: 291,  # Apartamento 175 -> Vaga 9
    372: 520,  # Apartamento 176 -> Vaga 238
    379: 284,  # Apartamento 184-1 -> Vaga 2
    380: 285,  # Apartamento 184-2 -> Vaga 3
    381: 535,  # Apartamento 185-1 -> Vaga 253
    382: 536,  # Apartamento 185-2 -> Vaga 254
    383: 533,  # Apartamento 186-1 -> Vaga 251
    384: 534,  # Apartamento 186-2 -> Vaga 252
    385: 296,  # Apartamento 2 -> Vaga 14
    386: 293,  # Apartamento 3 -> Vaga 11
    387: 342,  # Apartamento 4 -> Vaga 60
    393: 286,  # Apartamento 13 -> Vaga 4
    396: 419,  # Apartamento 16 -> Vaga 137
    398: 343,  # Apartamento 18 -> Vaga 61
    405: 315,  # Apartamento 27 -> Vaga 33
    409: 288,  # Apartamento 33 -> Vaga 6
    420: 298,  # Apartamento 46 -> Vaga 16
    427: 443,  # Apartamento 55 -> Vaga 161
    428: 416,  # Apartamento 56 -> Vaga 134
    430: 562,  # Apartamento 58 -> Vaga 280
    433: 319,  # Apartamento 63 -> Vaga 37
    434: 332,  # Apartamento 64 -> Vaga 50
    435: 444,  # Apartamento 65 -> Vaga 162
    439: 333,  # Apartamento 71 -> Vaga 51
    440: 320,  # Apartamento 72 -> Vaga 38
    442: 331,  # Apartamento 74 -> Vaga 49
    445: 412,  # Apartamento 77 -> Vaga 130
    446: 563,  # Apartamento 78 -> Vaga 281
    447: 560,  # Apartamento 81 -> Vaga 278
    449: 351,  # Apartamento 83 -> Vaga 69
    453: 431,  # Apartamento 87 -> Vaga 149
    456: 424,  # Apartamento 92 -> Vaga 142
    459: 283,  # Apartamento 95 -> Vaga 1
    464: 411,  # Apartamento 102 -> Vaga 129
    465: 427,  # Apartamento 103 -> Vaga 145
    468: 430,  # Apartamento 106 -> Vaga 148
    472: 425,  # Apartamento 112 -> Vaga 143
    473: 290,  # Apartamento 113 -> Vaga 8
    475: 558,  # Apartamento 115 -> Vaga 276
    481: 512,  # Apartamento 123 -> Vaga 230
    483: 499,  # Apartamento 125 -> Vaga 217
    486: 448,  # Apartamento 128 -> Vaga 166
    488: 420,  # Apartamento 132 -> Vaga 138
    490: 445,  # Apartamento 134 -> Vaga 163
    493: 330,  # Apartamento 137 -> Vaga 48
    505: 415,  # Apartamento 153 -> Vaga 133
    509: 426,  # Apartamento 157 -> Vaga 144
    515: 313,  # Apartamento 165 -> Vaga 31
    523: 359,  # Apartamento 175 -> Vaga 77
    528: 287,  # Apartamento 182 -> Vaga 5
    531: 418,  # Apartamento 185 -> Vaga 136
    532: 344,  # Apartamento 186 -> Vaga 62
    535: 379,  # Apartamento 191-1 -> Vaga 97
    536: 380,  # Apartamento 191-2 -> Vaga 98
    541: 396,  # Apartamento 194-1 -> Vaga 114
    542: 397,  # Apartamento 194-2 -> Vaga 115
    545: 381,  # Apartamento 196-1 -> Vaga 99
    546: 382,  # Apartamento 196-2 -> Vaga 100
    547: 377,  # Apartamento 197-1 -> Vaga 95
    548: 378,  # Apartamento 197-2 -> Vaga 96
}

# Create your views here.

# View para realizar o sorteio
def fatto_passion_sorteio(request):
    if request.method == 'POST':
        start_time = time.time()
        
        # Limpar registros anteriores de sorteio
        logger.info("Iniciando deleção dos registros anteriores")
        delete_start = time.time()
        Sorteio.objects.all().delete()
        logger.info(f"Deleção concluída em {time.time() - delete_start:.2f} segundos")

        # Obter todos os apartamentos e vagas
        logger.info("Buscando apartamentos e vagas")
        fetch_start = time.time()
        apartamentos = list(Apartamento.objects.all())
        vagas = list(Vaga.objects.all())
        logger.info(f"Busca concluída em {time.time() - fetch_start:.2f} segundos")

        # Lista para armazenar todos os objetos Sorteio
        sorteios_para_criar = []

        # FASE 1: Vagas Travadas (PNE e fixas)
        logger.info("FASE 1: Iniciando atribuição de vagas travadas")
        pne_start = time.time()
        
        for apartamento_id, vaga_id in APARTAMENTOS_VAGAS_FIXAS.items():
            try:
                apartamento = Apartamento.objects.get(id=apartamento_id)
                vaga = Vaga.objects.get(id=vaga_id)
                sorteios_para_criar.append(
                    Sorteio(
                        apartamento=apartamento,
                        vaga=vaga
                    )
                )
                # Remover o apartamento e a vaga das listas de sorteio
                apartamentos = [apt for apt in apartamentos if apt.id != apartamento_id]
                vagas = [v for v in vagas if v.id != vaga_id]
                logger.info(f"Vaga travada atribuída: Apartamento {apartamento_id} -> Vaga {vaga_id}")
            except (Apartamento.DoesNotExist, Vaga.DoesNotExist) as e:
                logger.error(f"Erro ao atribuir vaga travada: {e}")

        logger.info(f"FASE 1 concluída em {time.time() - pne_start:.2f} segundos")

        # FASE 2: Apartamentos com vagas descobertas
        logger.info("FASE 2: Sorteio de apartamentos com vagas descobertas")
        descoberta_start = time.time()
        
        # Filtrar apartamentos que precisam de vagas descobertas
        apartamentos_descobertas = [apt for apt in apartamentos if not apt.vaga_coberta]
        vagas_descobertas = [v for v in vagas if not v.vaga_coberta]
        
        if apartamentos_descobertas and vagas_descobertas:
            # Embaralhar apartamentos e vagas descobertas
            random.shuffle(apartamentos_descobertas)
            random.shuffle(vagas_descobertas)
            
            # Alocar vagas descobertas
            i = 0
            while i < len(apartamentos_descobertas) and i < len(vagas_descobertas):
                apartamento = apartamentos_descobertas[i]
                vaga = vagas_descobertas[i]
                
                sorteios_para_criar.append(
                    Sorteio(
                        apartamento=apartamento,
                        vaga=vaga
                    )
                )
                
                # Remover da lista geral
                apartamentos = [apt for apt in apartamentos if apt.id != apartamento.id]
                vagas = [v for v in vagas if v.id != vaga.id]
                i += 1

        logger.info(f"FASE 2 concluída em {time.time() - descoberta_start:.2f} segundos")

        # FASE 3: Subsolo 3 para apartamentos do bloco Vivant
        logger.info("FASE 3: Subsolo 3 para apartamentos do bloco Vivant")
        subsolo3_start = time.time()
        
        # Filtrar apartamentos do bloco Vivant
        apartamentos_vivant = [apt for apt in apartamentos if apt.bloco == '02.Vivant']
        vagas_subsolo3 = [v for v in vagas if v.andar == 'SUBSOLO_3']
        
        if apartamentos_vivant and vagas_subsolo3:
            # Embaralhar apartamentos Vivant e vagas subsolo 3
            random.shuffle(apartamentos_vivant)
            random.shuffle(vagas_subsolo3)
            
            # Alocar vagas subsolo 3 para apartamentos Vivant
            i = 0
            while i < len(apartamentos_vivant) and i < len(vagas_subsolo3):
                apartamento = apartamentos_vivant[i]
                vaga = vagas_subsolo3[i]
                
                sorteios_para_criar.append(
                    Sorteio(
                        apartamento=apartamento,
                        vaga=vaga
                    )
                )
                
                # Remover da lista geral
                apartamentos = [apt for apt in apartamentos if apt.id != apartamento.id]
                vagas = [v for v in vagas if v.id != vaga.id]
                i += 1

        logger.info(f"FASE 3 concluída em {time.time() - subsolo3_start:.2f} segundos")

        # FASE 4: Subsolo 1 para apartamentos do bloco Amour
        logger.info("FASE 4: Subsolo 1 para apartamentos do bloco Amour")
        subsolo1_start = time.time()
        
        # Filtrar apartamentos do bloco Amour
        apartamentos_amour = [apt for apt in apartamentos if apt.bloco == '01.Amour']
        vagas_subsolo1 = [v for v in vagas if v.andar == 'SUBSOLO_1']
        
        if apartamentos_amour and vagas_subsolo1:
            # Embaralhar apartamentos Amour e vagas subsolo 1
            random.shuffle(apartamentos_amour)
            random.shuffle(vagas_subsolo1)
            
            # Alocar vagas subsolo 1 para apartamentos Amour
            i = 0
            while i < len(apartamentos_amour) and i < len(vagas_subsolo1):
                apartamento = apartamentos_amour[i]
                vaga = vagas_subsolo1[i]
                
                sorteios_para_criar.append(
                    Sorteio(
                        apartamento=apartamento,
                        vaga=vaga
                    )
                )
                
                # Remover da lista geral
                apartamentos = [apt for apt in apartamentos if apt.id != apartamento.id]
                vagas = [v for v in vagas if v.id != vaga.id]
                i += 1

        logger.info(f"FASE 4 concluída em {time.time() - subsolo1_start:.2f} segundos")

        # FASE 5: Subsolo 2 misturado (ambos os blocos)
        logger.info("FASE 5: Subsolo 2 misturado para apartamentos restantes")
        subsolo2_start = time.time()
        
        # Filtrar vagas subsolo 2
        vagas_subsolo2 = [v for v in vagas if v.andar == 'SUBSOLO_2']
        
        if apartamentos and vagas_subsolo2:
            # Embaralhar apartamentos restantes e vagas subsolo 2
            random.shuffle(apartamentos)
            random.shuffle(vagas_subsolo2)
            
            # Alocar vagas subsolo 2 para apartamentos restantes
            i = 0
            while i < len(apartamentos) and i < len(vagas_subsolo2):
                apartamento = apartamentos[i]
                vaga = vagas_subsolo2[i]
                
                sorteios_para_criar.append(
                    Sorteio(
                        apartamento=apartamento,
                        vaga=vaga
                    )
                )
                
                i += 1

        logger.info(f"FASE 5 concluída em {time.time() - subsolo2_start:.2f} segundos")

        # Criar todos os registros de uma vez usando bulk_create
        Sorteio.objects.bulk_create(sorteios_para_criar)
        logger.info(f"Criação dos registros concluída em {time.time() - start_time:.2f} segundos")

        # Marcar o sorteio como iniciado e armazenar o horário de conclusão
        request.session['sorteio_iniciado'] = True
        request.session['horario_conclusao'] = timezone.localtime().strftime("%d/%m/%Y às %Hh %Mmin e %Ss")

        logger.info(f"Sorteio total concluído em {time.time() - start_time:.2f} segundos")
        
        # Redireciona para a mesma página após o sorteio
        return redirect(reverse('fatto_passion_sorteio'))

    # Se o método for GET, exibe os resultados ou a interface de sorteio
    sorteio_iniciado = request.session.get('sorteio_iniciado', False)
    vagas_atribuidas = Sorteio.objects.exists()
    resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id') if vagas_atribuidas else None

    context = {
        'sorteio_iniciado': sorteio_iniciado,
        'vagas_atribuidas': vagas_atribuidas,
        'resultados_sorteio': resultados_sorteio,
        'horario_conclusao': request.session.get('horario_conclusao', '')
    }

    return render(request, 'fatto_passion/fatto_passion_sorteio.html', context)

def fatto_passion_excel(request):
    # Caminho do modelo Excel
    caminho_modelo = 'setup/static/assets/modelos/sorteiofatto_passion.xlsx'

    # Carregar o modelo existente
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Ordenar os resultados do sorteio pelo ID do apartamento
    resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all()

    # Pegar o horário de conclusão do sorteio
    horario_conclusao = request.session.get('horario_conclusao', 'Horário não disponível')
    ws['B8'] = f"Sorteio realizado em: {horario_conclusao}"

    # Começar a partir da linha 10 (baseado no layout do seu modelo)
    linha = 11
    for sorteio in resultados_sorteio:
        ws[f'A{linha}'] = sorteio.apartamento.bloco
        ws[f'B{linha}'] = sorteio.apartamento.numero
        ws[f'C{linha}'] = sorteio.vaga.numero
        ws[f'D{linha}'] = sorteio.vaga.get_andar_display()
        ws[f'E{linha}'] = "Coberta" if sorteio.vaga.vaga_coberta else "Descoberta"
        ws[f'F{linha}'] = "PNE" if sorteio.apartamento.pne else "-"
        linha += 1

    # Configurar a resposta para o download do Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sorteio_fatto_passion.xlsx"'

    # Salvar o arquivo Excel na resposta
    wb.save(response)

    return response

def fatto_passion_qrcode(request):
    # Obter todos os apartamentos para preencher o dropdown
    apartamentos_disponiveis = Apartamento.objects.all().order_by('bloco', 'numero')
    
    # Obter o apartamento e bloco selecionados através do filtro (via GET)
    numero_apartamento = request.GET.get('apartamento')
    bloco_selecionado = request.GET.get('bloco')

    # Inicializar a variável de resultados filtrados como uma lista vazia
    resultados_filtrados = []

    # Se o apartamento foi selecionado, realizar a filtragem dos resultados do sorteio
    if numero_apartamento:
        # Construir o filtro base com o número do apartamento
        filtro = {'apartamento__numero': numero_apartamento}
        
        # Se um bloco específico foi selecionado (e não for "todos"), adicionar ao filtro
        if bloco_selecionado and bloco_selecionado != 'todos':
            filtro['apartamento__bloco'] = bloco_selecionado
        # Se "todos" foi selecionado, mas um apartamento específico foi escolhido,
        # precisamos determinar o bloco correto do apartamento selecionado
        elif bloco_selecionado == 'todos':
            # Buscar o apartamento específico para determinar seu bloco
            apartamento_especifico = Apartamento.objects.filter(numero=numero_apartamento).first()
            if apartamento_especifico:
                filtro['apartamento__bloco'] = apartamento_especifico.bloco
            
        # Buscar os sorteios com os filtros aplicados
        resultados_filtrados = list(Sorteio.objects.filter(**filtro).select_related('apartamento', 'vaga'))
        
        # Debug: verificar se há dados
        print(f"Filtro aplicado: {filtro}")
        print(f"Resultados encontrados: {len(resultados_filtrados)}")
        for resultado in resultados_filtrados:
            print(f"Apartamento: {resultado.apartamento.numero} - Bloco: {resultado.apartamento.bloco}")
            print(f"Vaga: {resultado.vaga.numero} - Andar: {resultado.vaga.get_andar_display()}")

    return render(request, 'fatto_passion/fatto_passion_qrcode.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': numero_apartamento,
        'bloco_selecionado': bloco_selecionado,
        'apartamentos_disponiveis': apartamentos_disponiveis,
    })

def fatto_passion_zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        return redirect('fatto_passion_sorteio')
    return render(request, 'fatto_passion/fatto_passion_zerar.html')

def fatto_passion_sorteio_v2(request):
    if request.method == 'POST':
        start_time = time.time()
        
        # Limpar registros anteriores de sorteio
        logger.info("Iniciando deleção dos registros anteriores")
        delete_start = time.time()
        Sorteio.objects.all().delete()
        logger.info(f"Deleção concluída em {time.time() - delete_start:.2f} segundos")

        # Obter todos os apartamentos e vagas
        logger.info("Buscando apartamentos e vagas")
        fetch_start = time.time()
        apartamentos = list(Apartamento.objects.all())
        vagas = list(Vaga.objects.all())
        logger.info(f"Busca concluída em {time.time() - fetch_start:.2f} segundos")

        # Lista para armazenar todos os objetos Sorteio
        sorteios_para_criar = []

        # FASE 1: Vagas Travadas (PNE e fixas) - VERSÃO V2
        logger.info("FASE 1: Iniciando atribuição de vagas travadas - VERSÃO V2")
        pne_start = time.time()
        
        # Aqui você pode definir manualmente as vagas específicas para a versão v2
        # Exemplo de como definir as vagas travadas manualmente:
        vagas_travadas_v2 = {
            # Apartamento ID: Vaga ID
            # Exemplo: 279: 413,  # Apartamento 23 -> Vaga 131
            # Adicione aqui as vagas específicas que você quer para a versão v2
            371: 291,  # Apartamento 175 -> Vaga 9
            386: 293,  # Apartamento 3 -> Vaga 11
            440: 320,  # Apartamento 72 -> Vaga 38
            493: 330,  # Apartamento 137 -> Vaga 48
            449: 351,  # Apartamento 83 -> Vaga 69
            314: 422,  # Apartamento 82 -> Vaga 140
            348: 516,  # Apartamento 136 -> Vaga 234
            320: 522,  # Apartamento 92 -> Vaga 240
            293: 538,  # Apartamento 45 -> Vaga 256
            283: 561,  # Apartamento 31 -> Vaga 279
            379: 284,  # Apartamento 184-1 -> Vaga 2
            380: 285,  # Apartamento 184-2 -> Vaga 3
        }
        
        for apartamento_id, vaga_id in vagas_travadas_v2.items():
            try:
                apartamento = Apartamento.objects.get(id=apartamento_id)
                vaga = Vaga.objects.get(id=vaga_id)
                sorteios_para_criar.append(
                    Sorteio(
                        apartamento=apartamento,
                        vaga=vaga
                    )
                )
                # Remover o apartamento e a vaga das listas de sorteio
                apartamentos = [apt for apt in apartamentos if apt.id != apartamento_id]
                vagas = [v for v in vagas if v.id != vaga_id]
                logger.info(f"Vaga travada V2 atribuída: Apartamento {apartamento_id} -> Vaga {vaga_id}")
            except (Apartamento.DoesNotExist, Vaga.DoesNotExist) as e:
                logger.error(f"Erro ao atribuir vaga travada V2: {e}")

        logger.info(f"FASE 1 V2 concluída em {time.time() - pne_start:.2f} segundos")

        # FASE 2: Apartamentos com vagas descobertas
        logger.info("FASE 2: Sorteio de apartamentos com vagas descobertas")
        descoberta_start = time.time()
        
        # Filtrar apartamentos que precisam de vagas descobertas
        apartamentos_descobertas = [apt for apt in apartamentos if not apt.vaga_coberta]
        vagas_descobertas = [v for v in vagas if not v.vaga_coberta]
        
        if apartamentos_descobertas and vagas_descobertas:
            # Embaralhar apartamentos e vagas descobertas
            random.shuffle(apartamentos_descobertas)
            random.shuffle(vagas_descobertas)
            
            # Alocar vagas descobertas
            i = 0
            while i < len(apartamentos_descobertas) and i < len(vagas_descobertas):
                apartamento = apartamentos_descobertas[i]
                vaga = vagas_descobertas[i]
                
                sorteios_para_criar.append(
                    Sorteio(
                        apartamento=apartamento,
                        vaga=vaga
                    )
                )
                
                # Remover da lista geral
                apartamentos = [apt for apt in apartamentos if apt.id != apartamento.id]
                vagas = [v for v in vagas if v.id != vaga.id]
                i += 1

        logger.info(f"FASE 2 concluída em {time.time() - descoberta_start:.2f} segundos")

        # FASE 3: Subsolo 3 para apartamentos do bloco Vivant
        logger.info("FASE 3: Subsolo 3 para apartamentos do bloco Vivant")
        subsolo3_start = time.time()
        
        # Filtrar apartamentos do bloco Vivant
        apartamentos_vivant = [apt for apt in apartamentos if apt.bloco == '02.Vivant']
        vagas_subsolo3 = [v for v in vagas if v.andar == 'SUBSOLO_3']
        
        if apartamentos_vivant and vagas_subsolo3:
            # Embaralhar apartamentos Vivant e vagas subsolo 3
            random.shuffle(apartamentos_vivant)
            random.shuffle(vagas_subsolo3)
            
            # Alocar vagas subsolo 3 para apartamentos Vivant
            i = 0
            while i < len(apartamentos_vivant) and i < len(vagas_subsolo3):
                apartamento = apartamentos_vivant[i]
                vaga = vagas_subsolo3[i]
                
                sorteios_para_criar.append(
                    Sorteio(
                        apartamento=apartamento,
                        vaga=vaga
                    )
                )
                
                # Remover da lista geral
                apartamentos = [apt for apt in apartamentos if apt.id != apartamento.id]
                vagas = [v for v in vagas if v.id != vaga.id]
                i += 1

        logger.info(f"FASE 3 concluída em {time.time() - subsolo3_start:.2f} segundos")

        # FASE 4: Subsolo 1 para apartamentos do bloco Amour
        logger.info("FASE 4: Subsolo 1 para apartamentos do bloco Amour")
        subsolo1_start = time.time()
        
        # Filtrar apartamentos do bloco Amour
        apartamentos_amour = [apt for apt in apartamentos if apt.bloco == '01.Amour']
        vagas_subsolo1 = [v for v in vagas if v.andar == 'SUBSOLO_1']
        
        if apartamentos_amour and vagas_subsolo1:
            # Embaralhar apartamentos Amour e vagas subsolo 1
            random.shuffle(apartamentos_amour)
            random.shuffle(vagas_subsolo1)
            
            # Alocar vagas subsolo 1 para apartamentos Amour
            i = 0
            while i < len(apartamentos_amour) and i < len(vagas_subsolo1):
                apartamento = apartamentos_amour[i]
                vaga = vagas_subsolo1[i]
                
                sorteios_para_criar.append(
                    Sorteio(
                        apartamento=apartamento,
                        vaga=vaga
                    )
                )
                
                # Remover da lista geral
                apartamentos = [apt for apt in apartamentos if apt.id != apartamento.id]
                vagas = [v for v in vagas if v.id != vaga.id]
                i += 1

        logger.info(f"FASE 4 concluída em {time.time() - subsolo1_start:.2f} segundos")

        # FASE 5: Subsolo 2 misturado (ambos os blocos)
        logger.info("FASE 5: Subsolo 2 misturado para apartamentos restantes")
        subsolo2_start = time.time()
        
        # Filtrar vagas subsolo 2
        vagas_subsolo2 = [v for v in vagas if v.andar == 'SUBSOLO_2']
        
        if apartamentos and vagas_subsolo2:
            # Embaralhar apartamentos restantes e vagas subsolo 2
            random.shuffle(apartamentos)
            random.shuffle(vagas_subsolo2)
            
            # Alocar vagas subsolo 2 para apartamentos restantes
            i = 0
            while i < len(apartamentos) and i < len(vagas_subsolo2):
                apartamento = apartamentos[i]
                vaga = vagas_subsolo2[i]
                
                sorteios_para_criar.append(
                    Sorteio(
                        apartamento=apartamento,
                        vaga=vaga
                    )
                )
                
                i += 1

        logger.info(f"FASE 5 concluída em {time.time() - subsolo2_start:.2f} segundos")

        # Criar todos os registros de uma vez usando bulk_create
        Sorteio.objects.bulk_create(sorteios_para_criar)
        logger.info(f"Criação dos registros concluída em {time.time() - start_time:.2f} segundos")

        # Marcar o sorteio como iniciado e armazenar o horário de conclusão
        request.session['sorteio_iniciado'] = True
        request.session['horario_conclusao'] = timezone.localtime().strftime("%d/%m/%Y às %Hh %Mmin e %Ss")

        logger.info(f"Sorteio V2 total concluído em {time.time() - start_time:.2f} segundos")
        
        # Redireciona para a mesma página após o sorteio
        return redirect(reverse('fatto_passion_sorteio'))

    # Se o método for GET, exibe os resultados ou a interface de sorteio
    sorteio_iniciado = request.session.get('sorteio_iniciado', False)
    vagas_atribuidas = Sorteio.objects.exists()
    resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id') if vagas_atribuidas else None

    context = {
        'sorteio_iniciado': sorteio_iniciado,
        'vagas_atribuidas': vagas_atribuidas,
        'resultados_sorteio': resultados_sorteio,
        'horario_conclusao': request.session.get('horario_conclusao', '')
    }

    return render(request, 'fatto_passion/fatto_passion_sorteio.html', context)
