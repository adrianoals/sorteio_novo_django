from django.shortcuts import render, redirect
from django.utils import timezone
import random
from django.http import HttpResponse
from django.urls import reverse
from openpyxl import load_workbook
from .models import Apartamento, Vaga, Sorteio

# View para realizar o sorteio
def la_corunha_sorteio(request):
    if request.method == 'POST':
        # Limpar registros anteriores de sorteio
        Sorteio.objects.all().delete()

        # Carregar apartamentos e vagas separados por tipo
        apts_carro = list(Apartamento.objects.filter(tipo_vaga_direito='Carro'))
        apts_moto = list(Apartamento.objects.filter(tipo_vaga_direito='Moto'))
        
        # Usar os campos booleanos para garantir filtro correto
        vagas_carro = list(Vaga.objects.filter(is_carro=True))
        vagas_moto = list(Vaga.objects.filter(is_moto=True))
        
        print(f"\n📊 DEBUG: Apartamentos Carro: {len(apts_carro)} | Vagas Carro: {len(vagas_carro)}")
        print(f"📊 DEBUG: Apartamentos Moto: {len(apts_moto)} | Vagas Moto: {len(vagas_moto)}")
        
        # Verificar se os números estão corretos
        if len(apts_carro) != 20:
            print(f"⚠️ ATENÇÃO: Esperado 20 apartamentos de Carro, encontrado {len(apts_carro)}")
        if len(apts_moto) != 9:
            print(f"⚠️ ATENÇÃO: Esperado 9 apartamentos de Moto, encontrado {len(apts_moto)}")
        if len(vagas_carro) != 14:
            print(f"⚠️ ATENÇÃO: Esperado 14 vagas de Carro, encontrado {len(vagas_carro)}")
        if len(vagas_moto) != 15:
            print(f"⚠️ ATENÇÃO: Esperado 15 vagas de Moto, encontrado {len(vagas_moto)}")

        # Usar set para rastrear IDs de vagas e apartamentos já atribuídos
        vagas_atribuidas_ids = set()
        apartamentos_atribuidos_ids = set()
        sorteios_para_criar = []

        # ============================================
        # 1. SORTEIO DE VAGAS DE CARRO
        # ============================================
        random.shuffle(apts_carro)
        random.shuffle(vagas_carro)
        
        # Garantir que temos vagas suficientes
        if len(apts_carro) > len(vagas_carro):
            print(f"⚠️ AVISO: {len(apts_carro)} apartamentos com direito a Carro, mas apenas {len(vagas_carro)} vagas disponíveis!")
        
        # Atribuir uma vaga para cada apartamento (até o limite de vagas disponíveis)
        for i, apt in enumerate(apts_carro):
            if i < len(vagas_carro):
                vaga_carro = vagas_carro[i]
                sorteios_para_criar.append(Sorteio(apartamento=apt, vaga=vaga_carro))
                vagas_atribuidas_ids.add(vaga_carro.id)
                apartamentos_atribuidos_ids.add(apt.id)
                print(f"✅ Apartamento {apt.numero} → Vaga Carro {vaga_carro.numero}")
            else:
                print(f"⚠️ Apartamento {apt.numero} não recebeu vaga (vagas insuficientes)")

        # ============================================
        # 2. SORTEIO DE VAGAS DE MOTO
        # ============================================
        random.shuffle(apts_moto)
        random.shuffle(vagas_moto)
        
        # Garantir que temos vagas suficientes
        if len(apts_moto) > len(vagas_moto):
            print(f"⚠️ AVISO: {len(apts_moto)} apartamentos com direito a Moto, mas apenas {len(vagas_moto)} vagas disponíveis!")
        
        # Atribuir uma vaga para cada apartamento (até o limite de vagas disponíveis)
        for i, apt in enumerate(apts_moto):
            if i < len(vagas_moto):
                vaga_moto = vagas_moto[i]
                sorteios_para_criar.append(Sorteio(apartamento=apt, vaga=vaga_moto))
                vagas_atribuidas_ids.add(vaga_moto.id)
                apartamentos_atribuidos_ids.add(apt.id)
                print(f"✅ Apartamento {apt.numero} → Vaga Moto {vaga_moto.numero}")
            else:
                print(f"⚠️ Apartamento {apt.numero} não recebeu vaga (vagas insuficientes)")

        # ============================================
        # 3. BULK CREATE (otimização de performance)
        # ============================================
        if sorteios_para_criar:
            Sorteio.objects.bulk_create(sorteios_para_criar)
            print(f"\n✅ Total de {len(sorteios_para_criar)} sorteios criados com sucesso!")

        # Marcar o sorteio como iniciado e armazenar o horário de conclusão
        request.session['sorteio_iniciado'] = True
        request.session['horario_conclusao'] = timezone.localtime().strftime("%d/%m/%Y às %Hh %Mmin e %Ss")

        # Redireciona para a mesma página após o sorteio
        return redirect(reverse('la_corunha_sorteio'))

    # Se o método for GET, exibe os resultados ou a interface de sorteio
    sorteio_iniciado = request.session.get('sorteio_iniciado', False)
    vagas_atribuidas = Sorteio.objects.exists()
    resultados_sorteio = Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id') if vagas_atribuidas else None

    context = {
        'sorteio_iniciado': sorteio_iniciado,
        'vagas_atribuidas': vagas_atribuidas,
        'resultados_sorteio': resultados_sorteio,
        'horario_conclusao': request.session.get('horario_conclusao', '')
    }

    return render(request, 'la_corunha/la_corunha_sorteio.html', context)


def la_corunha_excel(request):
    # Caminho do modelo Excel (você pode criar um modelo específico depois)
    caminho_modelo = 'setup/static/assets/modelos/sorteiolacorunha.xlsx' 

    # Carregar o modelo existente
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    # Pegar todos os sorteios ordenados por ID do apartamento
    resultados_sorteio = list(Sorteio.objects.select_related('apartamento', 'vaga').order_by('apartamento__id').all())

    # Pegar o horário de conclusão do sorteio
    horario_conclusao = request.session.get('horario_conclusao', 'Horário não disponível')
    ws['A8'] = f"Sorteio realizado em: {horario_conclusao}"

    # Limpar linhas antigas a partir da linha 10 até uma quantidade maior para garantir limpeza completa
    linha_inicial = 10
    linha_maxima = max(ws.max_row, linha_inicial + len(resultados_sorteio) + 50)  # Limpar mais linhas do que necessário
    
    for linha in range(linha_inicial, linha_maxima + 1):
        ws[f'A{linha}'] = None
        ws[f'B{linha}'] = None
        ws[f'C{linha}'] = None
        ws[f'D{linha}'] = None

    # Preencher dados
    linha = 10
    for sorteio in resultados_sorteio:
        ws[f'A{linha}'] = sorteio.apartamento.numero
        ws[f'B{linha}'] = sorteio.vaga.numero
        ws[f'C{linha}'] = sorteio.vaga.tipo_vaga
        linha += 1

    # Configurar a resposta para o download do Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sorteio_la_corunha.xlsx"'

    # Salvar o arquivo Excel na resposta
    wb.save(response)

    return response


def la_corunha_qrcode(request):
    # Obter o apartamento selecionado através do filtro (via GET)
    numero_apartamento = request.GET.get('apartamento')
    
    # Obter apartamentos que têm sorteio
    apartamentos_com_sorteio = Sorteio.objects.values_list('apartamento_id', flat=True).distinct()
    
    # Obter objetos Apartamento que têm sorteio (ordenados por ID)
    apartamentos_disponiveis = Apartamento.objects.filter(
        id__in=apartamentos_com_sorteio
    ).order_by('id')

    # Inicializar a variável de resultados filtrados como uma lista vazia
    resultados_filtrados = []

    # Se o apartamento foi selecionado, realizar a filtragem dos resultados do sorteio
    if numero_apartamento:
        # Buscar os sorteios para o apartamento selecionado com select_related para otimizar
        resultados_filtrados = Sorteio.objects.filter(
            apartamento__numero=numero_apartamento
        ).select_related('apartamento', 'vaga').order_by('apartamento__id')

    return render(request, 'la_corunha/la_corunha_qrcode.html', {
        'resultados_filtrados': resultados_filtrados,
        'apartamento_selecionado': numero_apartamento,
        'apartamentos_disponiveis': apartamentos_disponiveis,
    })


def la_corunha_zerar(request):
    if request.method == 'POST':
        Sorteio.objects.all().delete()
        request.session['sorteio_iniciado'] = False
        request.session.pop('horario_conclusao', None)
        return redirect('la_corunha_sorteio')
    else:
        return render(request, 'la_corunha/la_corunha_zerar.html')

